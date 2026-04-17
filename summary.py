from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from hpv_tdm import (
    AgeSexAggregateHPVModel,
    AgeSexSubtypeGroupedHPVModel,
    AggregateModelConfig,
    EvaluationConfig,
    EvaluationResult,
    Evaluator,
    SearchResult,
    SimulationResult,
    SubtypeGroupedModelConfig,
)
from hpv_tdm.result._plot import (
    PRODUCT_COLORS,
    apply_nature_style,
    apply_scientific_format,
)

BUDGET_BEGIN_YEAR = 2019
BUDGET_UNIT_LABEL = "10,000 yuan"
BUDGET_UNIT_DIVISOR = 10_000
DEFAULT_COPAY_PERSON_SHARES: tuple[float, float] = (0.6063, 0.2572)
TRIPARTY_MEDICAL_INSURANCE = 0.2791
TRIPARTY_GOVERNMENT = 0.2619
TRIPARTY_INDIVIDUAL = 0.4590
DEFAULT_FIG3_PRICE_DEMAND_METHOD = "weibull"
DEFAULT_FIG3_PRICE_DEMAND_COLUMN = "您预期的接种九价疫苗价格是多少全程3针"
DEFAULT_FIG3_PRICE_DEMAND_QUERY = "您对当前进口九价疫苗的看法全程3针==1"
DEFAULT_FIG3_PRICE_DEMAND_DOSES = 2
DEFAULT_FIG3_PRICE_DEMAND_WTP_SCALE = 0.5
PRICE_DEMAND_INCOME_COLUMN = "家庭月平均收入元"
PRICE_DEMAND_INCOME_LEVELS = np.array([1800, 6000, 9000, 11000, 18000, 23000, 35000])
PRICE_DEMAND_WTP_COLUMNS: dict[str, tuple[str, ...]] = {
    "domestic_bivalent": ("国产二价HPV疫苗愿意承受的最高价格是全程",),
    "imported_bivalent": ("您可接受的二价疫苗最高价格是多少全程3针",),
    "domestic_quadrivalent": ("国产四价HPV疫苗愿意承受的最高价格是全程",),
    "imported_quadrivalent": ("您可接受的四价疫苗最高价格是多少全程3针",),
    "domestic_nonavalent": (
        "国产宫九价HPV疫苗愿意承受的最高价格是全程",
        "国产九价HPV疫苗愿意承受的最高价格是全程",
    ),
    "imported_nonavalent": ("您可接受的九价疫苗最高价格是多少全程3针",),
}


@dataclass(frozen=True)
class SensitivityParameter:
    runtime_key: str
    table_label: str
    plot_label: str
    change_display: str
    value_unit: str


@dataclass(frozen=True)
class Fig2Scenario:
    years: int
    directory: Path
    time: np.ndarray
    age_labels: list[str]
    incidence_reduction: np.ndarray
    mortality_reduction: np.ndarray
    avoid_cecx: np.ndarray
    avoid_cecx_deaths: np.ndarray
    avoid_daly: np.ndarray
    vaccine_number: np.ndarray
    vaccine_cost: np.ndarray
    cost_saved: np.ndarray
    net_cost: np.ndarray
    ic_per_cecx: np.ndarray
    ic_per_cecx_death: np.ndarray
    icur: np.ndarray


SENSITIVITY_PARAMETERS: tuple[SensitivityParameter, ...] = (
    SensitivityParameter(
        runtime_key="epsilon_f",
        table_label=(
            "The probability that women are infected when they come into contact "
            "with HPV partners"
        ),
        plot_label="Female infection\nprob.",
        change_display="0.05",
        value_unit="plain",
    ),
    SensitivityParameter(
        runtime_key="epsilon_m",
        table_label="The probability that a man will be infected by an HPV partner",
        plot_label="Male infection\nprob.",
        change_display="0.05",
        value_unit="plain",
    ),
    SensitivityParameter(
        runtime_key="discount_rate",
        table_label="Bank rate",
        plot_label="Bank rate",
        change_display="0.02",
        value_unit="plain",
    ),
    SensitivityParameter(
        runtime_key="dose_cost",
        table_label="Per capita cost of vaccine procurement",
        plot_label="Vaccine cost",
        change_display="0.2",
        value_unit="rmb",
    ),
    SensitivityParameter(
        runtime_key="dose_cost",
        table_label=(
            "Per capita transportation and management costs and per capita "
            "service costs"
        ),
        plot_label="Service cost",
        change_display="0.2",
        value_unit="rmb",
    ),
    SensitivityParameter(
        runtime_key="cost_per_cecx",
        table_label="Cost of cervical cancer treatment (lifetime)",
        plot_label="CC treatment\ncost",
        change_display="0.2",
        value_unit="rmb",
    ),
    SensitivityParameter(
        runtime_key="daly_fatal",
        table_label="DALYs for cancer diagnosis",
        plot_label="Diagnosis\nDALY",
        change_display="0.1",
        value_unit="plain",
    ),
    SensitivityParameter(
        runtime_key="daly_fatal",
        table_label="DALYs for advanced cancer",
        plot_label="Advanced\nDALY",
        change_display="0.1",
        value_unit="plain",
    ),
)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate paper-ready summary tables and figures."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    tabs1_parser = subparsers.add_parser(
        "tabs1",
        help="Generate Supplementary Table s1 from the selected model configuration.",
    )
    tabs1_parser.add_argument(
        "--model-config",
        default="results/find_params/model_config_with_init_state.json",
        help="Model configuration file used to derive Supplementary Table s1.",
    )
    tabs1_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated summary files.",
    )

    sensitivity_parser = subparsers.add_parser(
        "sensitivity",
        help="Compute and persist ICUR one-way sensitivity analysis payloads.",
    )
    sensitivity_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    sensitivity_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    sensitivity_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated sensitivity payload files.",
    )

    budget_parser = subparsers.add_parser(
        "budget",
        help="Compute budget impact analysis and write a multi-sheet workbook.",
    )
    budget_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    budget_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    budget_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated budget workbook.",
    )

    copay_parser = subparsers.add_parser(
        "copay",
        help="Compute co-payment impact analysis and write a multi-sheet workbook.",
    )
    copay_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    copay_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    copay_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated co-payment workbook.",
    )
    copay_parser.add_argument(
        "--copay-person-shares",
        type=float,
        nargs=2,
        metavar=("SHARE1", "SHARE2"),
        default=DEFAULT_COPAY_PERSON_SHARES,
        help=(
            "Two personal payment shares used in the co-payment analysis. "
            "Defaults to 0.6063 and 0.2572."
        ),
    )

    demand_parser = subparsers.add_parser(
        "price-demand",
        help=(
            "Predict demand from the fitted price-demand curve for one "
            "income-price pair."
        ),
    )
    demand_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    demand_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    demand_parser.add_argument(
        "--price-demand-path",
        default="data/price_demand_data.dta",
        help="Path to the Stata file used for the price-demand function.",
    )
    demand_parser.add_argument(
        "--price-demand-method",
        choices=("weibull", "empirical"),
        default="weibull",
        help="Method used to estimate demand from the price-demand data.",
    )
    demand_parser.add_argument(
        "--price-demand-column",
        default=None,
        help=(
            "Optional explicit Stata column name used to build the price-demand "
            "curve. When omitted, the script chooses a default column based on "
            "the selected vaccine product."
        ),
    )
    demand_parser.add_argument(
        "--price-demand-query",
        default=None,
        help=(
            "Optional pandas query expression applied to the price-demand Stata "
            "dataset before fitting or plotting. Defaults to using all samples."
        ),
    )
    demand_parser.add_argument(
        "--price-demand-doses",
        type=int,
        choices=(2, 3),
        default=None,
        help=(
            "Optional dose count used to rescale the price-demand curve after "
            "converting the survey price to a per-dose basis. Defaults to the "
            "dose count used by the selected optimal strategy."
        ),
    )
    demand_parser.add_argument(
        "--price-demand-wtp-scale",
        type=float,
        default=1.0,
        help=(
            "Multiplier applied to WTP after converting the survey price to the "
            "chosen plotting dose schedule. Used to reflect the gap between "
            "stated willingness and actual behavior."
        ),
    )
    demand_parser.add_argument(
        "--income",
        type=float,
        required=True,
        help="Household monthly income used to predict demand.",
    )
    demand_target_group = demand_parser.add_mutually_exclusive_group(required=True)
    demand_target_group.add_argument(
        "--price",
        type=float,
        help="Vaccine price under the plotted dose schedule used to predict demand.",
    )
    demand_target_group.add_argument(
        "--demand",
        type=float,
        help="Target demand used to infer the corresponding vaccine price.",
    )

    triparty_parser = subparsers.add_parser(
        "triparty",
        help="Compute multi-party funding analysis and write a multi-sheet workbook.",
    )
    triparty_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    triparty_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    triparty_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated multi-party workbook.",
    )

    all_parser = subparsers.add_parser(
        "all",
        help="Run all summary subcommands in dependency-aware order.",
    )
    all_parser.add_argument(
        "--results-glob",
        default="search-*y",
        help="Glob pattern under results/ used to discover scenario directories.",
    )
    all_parser.add_argument(
        "--results-root",
        default="results",
        help="Root directory containing search result folders.",
    )
    all_parser.add_argument(
        "--output-dir",
        default="summary",
        help="Directory used to store the generated summary files.",
    )
    all_parser.add_argument(
        "--model-config",
        default="results/find_params/model_config_with_init_state.json",
        help="Model configuration file used to derive Supplementary Table s1.",
    )
    all_parser.add_argument(
        "--price-demand-path",
        default="data/price_demand_data.dta",
        help=(
            "Path to the Stata file used for the price-demand function. "
            "When unavailable, panel A is rendered as a placeholder."
        ),
    )
    all_parser.add_argument(
        "--price-demand-method",
        choices=("weibull", "empirical"),
        default=DEFAULT_FIG3_PRICE_DEMAND_METHOD,
        help=(
            "Method used to build the price-demand panel: a Weibull AFT fit or "
            "an empirical demand curve grouped by household income."
        ),
    )
    all_parser.add_argument(
        "--price-demand-column",
        default=DEFAULT_FIG3_PRICE_DEMAND_COLUMN,
        help=(
            "Optional explicit Stata column name used to build the price-demand "
            "curve. When omitted, the script chooses a default column based on "
            "the selected vaccine product."
        ),
    )
    all_parser.add_argument(
        "--price-demand-query",
        default=DEFAULT_FIG3_PRICE_DEMAND_QUERY,
        help=(
            "Optional pandas query expression applied to the price-demand Stata "
            "dataset before fitting or plotting. Defaults to using all samples."
        ),
    )
    all_parser.add_argument(
        "--price-demand-doses",
        type=int,
        choices=(2, 3),
        default=DEFAULT_FIG3_PRICE_DEMAND_DOSES,
        help=(
            "Optional dose count used to rescale the price-demand curve after "
            "converting the survey price to a per-dose basis. Defaults to the "
            "dose count used by the selected optimal strategy."
        ),
    )
    all_parser.add_argument(
        "--price-demand-wtp-scale",
        type=float,
        default=DEFAULT_FIG3_PRICE_DEMAND_WTP_SCALE,
        help=(
            "Multiplier applied to WTP after converting the survey price to the "
            "chosen plotting dose schedule. Used to reflect the gap between "
            "stated willingness and actual behavior."
        ),
    )
    all_parser.add_argument(
        "--sensitivity-path",
        default=None,
        help=(
            "Optional path to a precomputed sensitivity payload JSON file. "
            "Defaults to <output-dir>/sensitivity_s3.json."
        ),
    )
    all_parser.add_argument(
        "--budget-path",
        default=None,
        help=(
            "Optional path to the budget workbook. "
            "Defaults to <output-dir>/budget_impact.xlsx."
        ),
    )
    all_parser.add_argument(
        "--copay-path",
        default=None,
        help=(
            "Optional path to the co-payment workbook. "
            "Defaults to <output-dir>/copay_impact.xlsx."
        ),
    )
    all_parser.add_argument(
        "--triparty-path",
        default=None,
        help=(
            "Optional path to the multi-party workbook. "
            "Defaults to <output-dir>/triparty_impact.xlsx."
        ),
    )
    all_parser.add_argument(
        "--copay-person-shares",
        type=float,
        nargs=2,
        metavar=("SHARE1", "SHARE2"),
        default=DEFAULT_COPAY_PERSON_SHARES,
        help=(
            "Two personal payment shares used in the co-payment analysis. "
            "Defaults to 0.6063 and 0.2572."
        ),
    )

    for command_name, help_text in (
        (
            "tab1",
            "Summarize the selected optimal vaccination strategies across scenarios.",
        ),
        (
            "fig2",
            "Generate the main-text Figure 2 for age-heterogeneous "
            "health and economic outcomes.",
        ),
        (
            "fig3",
            "Generate the main-text Figure 3 for financing and payment scenarios.",
        ),
        (
            "figs1",
            "Generate Supplementary Figure s1 by combining Pareto fronts "
            "across scenarios.",
        ),
        (
            "figs2",
            "Generate Supplementary Figure s2 by combining search-history "
            "panels across scenarios.",
        ),
        (
            "tabs3",
            "Generate Supplementary Table s3 with ICUR one-way sensitivity analysis.",
        ),
        (
            "figs3",
            "Generate Supplementary Figure s3 with ICUR tornado plots.",
        ),
        (
            "figs4",
            "Generate Supplementary Figure s4 by combining budget impact panels.",
        ),
        (
            "figs5",
            "Generate Supplementary Figure s5 by combining co-payment panels.",
        ),
        (
            "figs6",
            "Generate Supplementary Figure s6 by combining multi-party panels.",
        ),
    ):
        figure_parser = subparsers.add_parser(command_name, help=help_text)
        figure_parser.add_argument(
            "--results-glob",
            default="search-*y",
            help="Glob pattern under results/ used to discover scenario directories.",
        )
        figure_parser.add_argument(
            "--results-root",
            default="results",
            help="Root directory containing search result folders.",
        )
        figure_parser.add_argument(
            "--output-dir",
            default="summary",
            help="Directory used to store the generated summary files.",
        )
        if command_name in {"tabs3", "figs3"}:
            figure_parser.add_argument(
                "--sensitivity-path",
                default=None,
                help=(
                    "Path to a precomputed sensitivity payload JSON file. "
                    "Defaults to <output-dir>/sensitivity_s3.json."
                ),
            )
        if command_name == "figs4":
            figure_parser.add_argument(
                "--budget-path",
                default="summary/budget_impact.xlsx",
                help="Path to the budget impact workbook generated by `budget`.",
            )
        if command_name == "fig3":
            figure_parser.add_argument(
                "--budget-path",
                default="summary/budget_impact.xlsx",
                help="Path to the budget impact workbook generated by `budget`.",
            )
            figure_parser.add_argument(
                "--copay-path",
                default="summary/copay_impact.xlsx",
                help="Path to the co-payment workbook generated by `copay`.",
            )
            figure_parser.add_argument(
                "--triparty-path",
                default="summary/triparty_impact.xlsx",
                help="Path to the multi-party workbook generated by `triparty`.",
            )
            figure_parser.add_argument(
                "--price-demand-path",
                default="data/price_demand_data.dta",
                help=(
                    "Path to the Stata file used for the price-demand function. "
                    "When unavailable, panel A is rendered as a placeholder."
                ),
            )
            figure_parser.add_argument(
                "--price-demand-method",
                choices=("weibull", "empirical"),
                default=DEFAULT_FIG3_PRICE_DEMAND_METHOD,
                help=(
                    "Method used to build the price-demand panel: a Weibull AFT "
                    "fit or an empirical demand curve grouped by household income."
                ),
            )
            figure_parser.add_argument(
                "--price-demand-column",
                default=DEFAULT_FIG3_PRICE_DEMAND_COLUMN,
                help=(
                    "Optional explicit Stata column name used to build the "
                    "price-demand curve. When omitted, the script chooses a "
                    "default column based on the selected vaccine product."
                ),
            )
            figure_parser.add_argument(
                "--price-demand-query",
                default=DEFAULT_FIG3_PRICE_DEMAND_QUERY,
                help=(
                    "Optional pandas query expression applied to the "
                    "price-demand Stata dataset before fitting or plotting. "
                    "Defaults to using all samples."
                ),
            )
            figure_parser.add_argument(
                "--price-demand-doses",
                type=int,
                choices=(2, 3),
                default=DEFAULT_FIG3_PRICE_DEMAND_DOSES,
                help=(
                    "Optional dose count used to rescale the price-demand curve "
                    "after converting the survey price to a per-dose basis. "
                    "Defaults to the dose count used by the selected optimal "
                    "strategy."
                ),
            )
            figure_parser.add_argument(
                "--price-demand-wtp-scale",
                type=float,
                default=DEFAULT_FIG3_PRICE_DEMAND_WTP_SCALE,
                help=(
                    "Multiplier applied to WTP after converting the survey "
                    "price to the chosen plotting dose schedule. Used to "
                    "reflect the gap between stated willingness and actual "
                    "behavior."
                ),
            )
        if command_name == "figs5":
            figure_parser.add_argument(
                "--copay-path",
                default="summary/copay_impact.xlsx",
                help="Path to the co-payment workbook generated by `copay`.",
            )
        if command_name == "figs6":
            figure_parser.add_argument(
                "--triparty-path",
                default="summary/triparty_impact.xlsx",
                help="Path to the multi-party workbook generated by `triparty`.",
            )

    return parser.parse_args()


def _scenario_years_from_name(directory: Path) -> int:
    stem = directory.name
    match = re.match(r"^search-(\d+)y(?:-.*)?$", stem)
    if match is None:
        raise ValueError(f"unexpected scenario directory name: {directory}")
    return int(match.group(1))


def _age_span_label(age_span: str, agebins: list[float]) -> str:
    start_text, stop_text = age_span.split(":")
    start = int(start_text)
    stop = int(stop_text)
    lower = agebins[start]
    upper = agebins[stop + 1]
    lower_label = int(lower) if float(lower).is_integer() else lower
    upper_label = int(upper) if float(upper).is_integer() else upper
    return f"[{lower_label}, {upper_label})"


def _product_label(product_id: str) -> str:
    mapping = {
        "bivalent": "Bivalent vaccine",
        "quadrivalent": "Quadrivalent vaccine",
        "nonavalent": "Nonavalent vaccine",
        "domestic_bivalent": "Domestic bivalent vaccine",
        "imported_bivalent": "Imported bivalent vaccine",
        "imported_quadrivalent": "Imported quadrivalent vaccine",
        "domestic_nonavalent": "Domestic nonavalent vaccine",
        "imported_nonavalent": "Imported nonavalent vaccine",
    }
    return mapping.get(product_id, product_id)


def _format_percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def _adjust_excel_widths(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    for column_cells in worksheet.columns:
        text_lengths = [
            len(str(cell.value)) for cell in column_cells if cell.value is not None
        ]
        width = max(text_lengths, default=0) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(width, 12), 36
        )
    workbook.save(path)


def _load_model_config(path: Path) -> AggregateModelConfig | SubtypeGroupedModelConfig:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    model_kind = payload.get("model_kind", "aggregate")
    if model_kind == "aggregate":
        return AggregateModelConfig.from_json_dict(payload)
    if model_kind == "subtype_grouped":
        return SubtypeGroupedModelConfig.from_json_dict(payload)
    raise ValueError(f"unsupported model kind: {model_kind}")


def _load_evaluation_config(path: Path) -> EvaluationConfig:
    return EvaluationConfig.from_json_file(path)


def _build_model(
    config: AggregateModelConfig | SubtypeGroupedModelConfig,
) -> AgeSexAggregateHPVModel | AgeSexSubtypeGroupedHPVModel:
    if isinstance(config, AggregateModelConfig) and not isinstance(
        config,
        SubtypeGroupedModelConfig,
    ):
        return AgeSexAggregateHPVModel(config)
    return AgeSexSubtypeGroupedHPVModel(config)


def _format_table_value(value: float) -> str:
    if value == 0:
        return "0"
    return f"{value:.5f}".rstrip("0").rstrip(".")


def _table_age_label(lower: float, upper: float) -> str:
    lower_text = str(int(lower)) if float(lower).is_integer() else str(lower)
    if upper == float("inf"):
        return f"[{lower_text}, inf)"
    upper_text = str(int(upper)) if float(upper).is_integer() else str(upper)
    return f"[{lower_text}, {upper_text})"


def _build_tabs1_dataframe(
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
) -> pd.DataFrame:
    demography = model_config.demography
    transmission = model_config.transmission
    agebins = demography.agebins
    rows: list[dict[str, str]] = []
    for idx in range(len(agebins) - 1):
        rows.append(
            {
                "Age group": _table_age_label(agebins[idx], agebins[idx + 1]),
                "Number of sexual partners per year for women": _format_table_value(
                    transmission.omega_f[idx]
                ),
                "Number of sexual partners per year for men": _format_table_value(
                    transmission.omega_m[idx]
                ),
                "Mortality of local cancer": _format_table_value(transmission.dL[idx]),
                "Mortality rate of region cancer": _format_table_value(
                    transmission.dR[idx]
                ),
                "Mortality of distant cancer": _format_table_value(
                    transmission.dD[idx]
                ),
                "Fertility rate": _format_table_value(demography.fertilities[idx]),
                "Death rate of women": _format_table_value(
                    demography.deathes_female[idx]
                ),
                "Death rate of men": _format_table_value(demography.deathes_male[idx]),
            }
        )
    return pd.DataFrame.from_records(rows)


def _write_table_outputs(
    dataframe: pd.DataFrame, output_dir: Path, stem: str
) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / f"{stem}.xlsx"
    csv_path = output_dir / f"{stem}.csv"
    markdown_path = output_dir / f"{stem}.md"

    dataframe.to_excel(excel_path, index=False)
    dataframe.to_csv(csv_path, index=False)
    markdown_lines = []
    headers = list(dataframe.columns)
    markdown_lines.append("| " + " | ".join(headers) + " |")
    markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in dataframe.itertuples(index=False, name=None):
        markdown_lines.append("| " + " | ".join(str(value) for value in row) + " |")
    markdown_path.write_text("\n".join(markdown_lines) + "\n", encoding="utf-8")
    _adjust_excel_widths(excel_path)
    return excel_path, csv_path, markdown_path


def _load_tab1_row(search_dir: Path) -> dict[str, object]:
    strategy = _load_search_strategy_metadata(search_dir)
    evaluation = EvaluationResult.from_hdf(search_dir / "best_evaluation.h5")

    return {
        "Scenario": strategy["scenario"],
        "Time horizon (years)": strategy["years"],
        "Target age group": strategy["target_age_group"],
        "Vaccine type": strategy["vaccine_type"],
        "Coverage": strategy["coverage"],
        "ICUR": strategy["icur"],
        "Final incidence (/100k)": float(evaluation.incidence[-1] * 1e5),
        "Final mortality (/100k)": float(evaluation.mortality[-1] * 1e5),
        "Averted cervical cancer cases": float(evaluation.avoid_cecx.sum(axis=1)[-1]),
        "Averted cervical cancer deaths": float(
            evaluation.avoid_cecx_deaths.sum(axis=1)[-1]
        ),
        "Averted DALYs": float(evaluation.avoid_daly[-1]),
        "Vaccine cost (yuan)": float(evaluation.cost_vacc[-1]),
        "Cervical cancer treatment cost (yuan)": float(evaluation.cost_cecx[-1]),
        "Total cost (yuan)": float(evaluation.total_cost[-1]),
    }


def _build_tab1_dataframe(results_root: Path, results_glob: str) -> pd.DataFrame:
    search_dirs = sorted(
        results_root.glob(results_glob),
        key=_scenario_years_from_name,
    )
    if not search_dirs:
        raise ValueError(
            "no search directories matched pattern "
            f"{results_glob!r} under {results_root}"
        )
    records = [_load_tab1_row(directory) for directory in search_dirs]
    return pd.DataFrame.from_records(records)


def _discover_search_dirs(results_root: Path, results_glob: str) -> list[Path]:
    search_dirs = sorted(
        results_root.glob(results_glob),
        key=_scenario_years_from_name,
    )
    if not search_dirs:
        raise ValueError(
            "no search directories matched pattern "
            f"{results_glob!r} under {results_root}"
        )
    return search_dirs


def _format_copay_scheme_code(vacc_reimburse: float) -> str:
    return f"{int(round(vacc_reimburse * 10_000)):04d}"


def _build_copay_schemes(
    person_shares: tuple[float, float] | list[float],
) -> tuple[tuple[str, float, float], ...]:
    schemes: list[tuple[str, float, float]] = []
    for person_share in person_shares:
        if not 0 <= float(person_share) <= 1:
            raise ValueError(
                f"co-payment personal share must be within [0, 1], got {person_share}"
            )
        vacc_reimburse = 1.0 - float(person_share)
        schemes.append(
            (
                _format_copay_scheme_code(vacc_reimburse),
                vacc_reimburse,
                float(person_share),
            )
        )
    return tuple(schemes)


def _copay_scheme_title(scheme_code: str) -> str:
    medical_share = int(scheme_code) / 10_000
    personal_share = 1.0 - medical_share
    return (
        f"personal {personal_share * 100:.2f}% / "
        f"medical insurance {medical_share * 100:.2f}%"
    )


def _copay_scheme_meta_from_dataframe(
    dataframe: pd.DataFrame,
) -> tuple[tuple[str, str], ...]:
    scheme_codes = []
    for column in dataframe.columns:
        match = re.fullmatch(r"(\d{4}) Vaccination cost", str(column))
        if match is not None:
            scheme_codes.append(match.group(1))
    if not scheme_codes:
        raise ValueError("no co-payment scheme columns were found in the workbook")
    return tuple((code, _copay_scheme_title(code)) for code in sorted(scheme_codes))


def _load_search_strategy_metadata(search_dir: Path) -> dict[str, object]:
    with (search_dir / "best_trial.json").open("r", encoding="utf-8") as handle:
        best_trial = json.load(handle)
    with (search_dir / "model_config.json").open("r", encoding="utf-8") as handle:
        model_config = json.load(handle)

    years = _scenario_years_from_name(search_dir)
    agebins = model_config["demography"]["agebins"]
    target_product_id = best_trial["params"]["target_product_id"]
    target_age_span = best_trial["params"]["target_age_span"]
    coverage = float(best_trial["params"]["coverage"])
    product = model_config["vaccine_catalog"]["products"][target_product_id]

    return {
        "scenario": f"{years} Years",
        "years": years,
        "target_age_group": _age_span_label(target_age_span, agebins),
        "target_age_span": target_age_span,
        "vaccine_type": _product_label(target_product_id),
        "product_id": target_product_id,
        "coverage": _format_percent(coverage),
        "icur": float(best_trial["values"][0]),
        "dose_price": float(product["dose_cost"]),
        "dose_count": _dose_count_for_age_span(product, agebins, target_age_span),
        "full_course_price": _full_course_price(product, agebins, target_age_span),
    }


def _full_course_price(
    product: dict[str, object], agebins: list[float], target_age_span: str
) -> float:
    return float(product["dose_cost"]) * float(
        _dose_count_for_age_span(product, agebins, target_age_span)
    )


def _dose_count_for_age_span(
    product: dict[str, object], agebins: list[float], target_age_span: str
) -> int:
    start_text, stop_text = target_age_span.split(":")
    lower = agebins[int(start_text)]
    upper = agebins[int(stop_text) + 1]
    if not np.isfinite(upper):
        raise ValueError(
            f"target age span must have finite upper bound: {target_age_span}"
        )

    dose_counts = {
        _dose_count_for_age(product["dose_schedules"], age)
        for age in range(int(lower), int(upper))
    }
    if len(dose_counts) != 1:
        raise ValueError(
            "target age span crosses multiple dose schedules, cannot infer a single "
            f"full-course price: {target_age_span}"
        )
    return int(dose_counts.pop())


def _dose_count_for_age(dose_schedules: list[dict[str, object]], age: int) -> int:
    matched_doses = [
        int(rule["doses"])
        for rule in dose_schedules
        if int(rule["age_min"]) <= age <= int(rule["age_max"])
    ]
    if len(matched_doses) != 1:
        raise ValueError(f"age {age} matched {len(matched_doses)} dose schedules")
    return matched_doses[0]


def _validate_sensitivity_dir(search_dir: Path) -> None:
    required = (
        "best_model_config.json",
        "evaluation_config.json",
        "best_trial.json",
    )
    missing = [name for name in required if not (search_dir / name).exists()]
    if missing:
        raise ValueError(
            f"search result directory {search_dir} is missing required files: "
            f"{', '.join(missing)}"
        )


def _build_reference_config(
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
) -> AggregateModelConfig | SubtypeGroupedModelConfig:
    return model_config.with_vaccination(
        product_id=None,
        coverage_by_age=[0.0] * model_config.nages,
    )


def _compact_age_label(label: str) -> str:
    match = re.match(r"^\[(.+),\s*(.+)\)$", label.strip())
    if match is None:
        return label
    lower, upper = match.groups()
    if upper == "inf":
        return f"{lower}+"
    return f"{lower}\u2013{upper}"


def _rate_matrix_per_100k(
    matrix: np.ndarray,
    population: np.ndarray,
) -> np.ndarray:
    return (
        np.divide(
            matrix,
            population,
            out=np.zeros_like(matrix, dtype=float),
            where=population > 0,
        )
        * 1e5
    )


def _safe_ratio(
    numerator: np.ndarray,
    denominator: np.ndarray,
) -> np.ndarray:
    return np.divide(
        numerator,
        denominator,
        out=np.full_like(numerator, np.nan, dtype=float),
        where=denominator > 0,
    )


def _build_fig2_scenario(search_dir: Path) -> Fig2Scenario:
    required = (
        "best_model_config.json",
        "evaluation_config.json",
        "best_simulation.h5",
    )
    missing = [name for name in required if not (search_dir / name).exists()]
    if missing:
        raise ValueError(
            f"search result directory {search_dir} is missing required files: "
            f"{', '.join(missing)}"
        )

    years = _scenario_years_from_name(search_dir)
    model_config = _load_model_config(search_dir / "best_model_config.json")
    evaluation_config = _load_evaluation_config(search_dir / "evaluation_config.json")

    vaccinated_simulation = SimulationResult.from_hdf(search_dir / "best_simulation.h5")
    vaccinated_model = vaccinated_simulation.get_model()
    reference_model = _build_model(_build_reference_config(model_config))
    reference_simulation = reference_model.simulate()

    evaluator = Evaluator(evaluation_config)
    evaluation = evaluator.evaluate(vaccinated_simulation, reference_simulation)
    vaccinated_absolute = evaluator._evaluate_absolute(vaccinated_simulation)
    reference_absolute = evaluator._evaluate_absolute(reference_simulation)

    vaccinated_population = np.asarray(
        vaccinated_model.total_female_population(vaccinated_simulation.state),
        dtype=float,
    )
    reference_population = np.asarray(
        reference_model.total_female_population(reference_simulation.state),
        dtype=float,
    )
    vaccinated_incidence = _rate_matrix_per_100k(
        np.asarray(vaccinated_model.incidence_matrix(vaccinated_simulation.state)),
        vaccinated_population,
    )
    reference_incidence = _rate_matrix_per_100k(
        np.asarray(reference_model.incidence_matrix(reference_simulation.state)),
        reference_population,
    )
    vaccinated_mortality = _rate_matrix_per_100k(
        np.asarray(vaccinated_model.mortality_matrix(vaccinated_simulation.state)),
        vaccinated_population,
    )
    reference_mortality = _rate_matrix_per_100k(
        np.asarray(reference_model.mortality_matrix(reference_simulation.state)),
        reference_population,
    )

    incidence_reduction = np.clip(reference_incidence - vaccinated_incidence, 0, None)
    mortality_reduction = np.clip(reference_mortality - vaccinated_mortality, 0, None)

    total_cost_diff = vaccinated_absolute.total_cost - reference_absolute.total_cost
    avoid_cecx = np.asarray(evaluation.avoid_cecx.sum(axis=1), dtype=float)
    avoid_cecx_deaths = np.asarray(
        evaluation.avoid_cecx_deaths.sum(axis=1),
        dtype=float,
    )
    avoid_daly = np.asarray(evaluation.avoid_daly, dtype=float)
    vaccine_number = (
        np.asarray(vaccinated_absolute.cumulative_vaccinated.sum(axis=1), dtype=float)
        / 10_000
    )
    vaccine_cost = np.asarray(vaccinated_absolute.cost_vacc, dtype=float) / (
        BUDGET_UNIT_DIVISOR
    )
    cost_saved = (
        np.asarray(reference_absolute.cost_cecx, dtype=float)
        - np.asarray(vaccinated_absolute.cost_cecx, dtype=float)
    ) / BUDGET_UNIT_DIVISOR
    net_cost = vaccine_cost - cost_saved

    return Fig2Scenario(
        years=years,
        directory=search_dir,
        time=np.asarray(vaccinated_simulation.time, dtype=float),
        age_labels=[
            _compact_age_label(label) for label in vaccinated_model.agebin_names
        ],
        incidence_reduction=incidence_reduction,
        mortality_reduction=mortality_reduction,
        avoid_cecx=avoid_cecx,
        avoid_cecx_deaths=avoid_cecx_deaths,
        avoid_daly=avoid_daly,
        vaccine_number=vaccine_number,
        vaccine_cost=vaccine_cost,
        cost_saved=cost_saved,
        net_cost=net_cost,
        ic_per_cecx=_safe_ratio(total_cost_diff, avoid_cecx),
        ic_per_cecx_death=_safe_ratio(total_cost_diff, avoid_cecx_deaths),
        icur=np.asarray(evaluation.icur, dtype=float),
    )


def _replace_transmission_value(
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
    key: str,
    value: float,
) -> AggregateModelConfig | SubtypeGroupedModelConfig:
    transmission = model_config.transmission.model_copy(update={key: value})
    return model_config.model_copy(update={"transmission": transmission})


def _replace_product_dose_cost(
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
    dose_cost: float,
) -> AggregateModelConfig | SubtypeGroupedModelConfig:
    product_id = model_config.resolved_product_id()
    product = model_config.vaccine_catalog.get_product(product_id)
    products = dict(model_config.vaccine_catalog.products)
    products[product_id] = product.model_copy(update={"dose_cost": dose_cost})
    vaccine_catalog = model_config.vaccine_catalog.model_copy(
        update={"products": products}
    )
    return model_config.model_copy(update={"vaccine_catalog": vaccine_catalog})


def _replace_evaluation_value(
    evaluation_config: EvaluationConfig,
    key: str,
    value: float,
) -> EvaluationConfig:
    return evaluation_config.model_copy(update={key: value})


def _format_numeric(value: float, decimals: int = 2) -> str:
    return f"{value:.{decimals}f}"


def _simulate_and_evaluate(
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
    evaluation_config: EvaluationConfig,
) -> float:
    vaccinated_model = _build_model(model_config)
    vaccinated_simulation = vaccinated_model.simulate()
    reference_model = _build_model(_build_reference_config(model_config))
    reference_simulation = reference_model.simulate()
    evaluator = Evaluator(evaluation_config)
    evaluation = evaluator.evaluate(vaccinated_simulation, reference_simulation)
    return float(evaluation.icur[-1])


def _evaluate_from_existing_simulations(
    vaccinated_simulation,
    reference_simulation,
    evaluation_config: EvaluationConfig,
) -> float:
    evaluator = Evaluator(evaluation_config)
    evaluation = evaluator.evaluate(vaccinated_simulation, reference_simulation)
    return float(evaluation.icur[-1])


def _parameter_display_values(
    parameter: SensitivityParameter,
    model_config: AggregateModelConfig | SubtypeGroupedModelConfig,
    evaluation_config: EvaluationConfig,
) -> tuple[float, float, float]:
    if parameter.runtime_key == "epsilon_f":
        original = float(model_config.transmission.epsilon_f)
        return original, original * 0.95, original * 1.05
    if parameter.runtime_key == "epsilon_m":
        original = float(model_config.transmission.epsilon_m)
        return original, original * 0.95, original * 1.05
    if parameter.runtime_key == "discount_rate":
        original = float(evaluation_config.discount_rate)
        return original, 0.01, 0.05
    if parameter.runtime_key == "dose_cost":
        original = float(
            model_config.vaccine_catalog.get_product(
                model_config.resolved_product_id()
            ).dose_cost
        )
        return original, original * 0.8, original * 1.2
    if parameter.runtime_key == "cost_per_cecx":
        original = float(evaluation_config.cost_per_cecx)
        return original, original * 0.8, original * 1.2
    if parameter.runtime_key == "daly_fatal":
        original = float(evaluation_config.daly_fatal)
        return original, original * 0.9, original * 1.1
    raise ValueError(f"unsupported sensitivity parameter: {parameter.runtime_key}")


def _format_parameter_value(value: float, unit: str) -> str:
    display = value
    if unit == "plain" and abs(display) < 1:
        return f"{display:.4f}".rstrip("0").rstrip(".")
    return _format_numeric(display, decimals=2)


def _compute_sensitivity_payloads(
    results_root: Path,
    results_glob: str,
) -> tuple[list[tuple[int, Path]], list[dict[str, object]]]:
    search_dirs = _discover_search_dirs(results_root, results_glob)
    scenarios = [
        (_scenario_years_from_name(directory), directory) for directory in search_dirs
    ]
    rows: list[dict[str, object]] = []
    for years, search_dir in scenarios:
        _validate_sensitivity_dir(search_dir)
        model_config = _load_model_config(search_dir / "best_model_config.json")
        evaluation_config = _load_evaluation_config(
            search_dir / "evaluation_config.json"
        )

        vaccinated_model = _build_model(model_config)
        vaccinated_simulation = vaccinated_model.simulate()
        reference_model = _build_model(_build_reference_config(model_config))
        reference_simulation = reference_model.simulate()
        baseline_icur = _evaluate_from_existing_simulations(
            vaccinated_simulation,
            reference_simulation,
            evaluation_config,
        )

        runtime_results: dict[str, dict[str, float]] = {
            "baseline": {"original": baseline_icur}
        }

        epsilon_f = float(model_config.transmission.epsilon_f)
        epsilon_f_results = {}
        for bound_name, bound_value in {
            "lower": epsilon_f * 0.95,
            "upper": epsilon_f * 1.05,
        }.items():
            perturbed_model_config = _replace_transmission_value(
                model_config,
                "epsilon_f",
                bound_value,
            )
            epsilon_f_results[bound_name] = _simulate_and_evaluate(
                perturbed_model_config,
                evaluation_config,
            )
        runtime_results["epsilon_f"] = epsilon_f_results

        epsilon_m = float(model_config.transmission.epsilon_m)
        epsilon_m_results = {}
        for bound_name, bound_value in {
            "lower": epsilon_m * 0.95,
            "upper": epsilon_m * 1.05,
        }.items():
            perturbed_model_config = _replace_transmission_value(
                model_config,
                "epsilon_m",
                bound_value,
            )
            epsilon_m_results[bound_name] = _simulate_and_evaluate(
                perturbed_model_config,
                evaluation_config,
            )
        runtime_results["epsilon_m"] = epsilon_m_results

        dose_cost = float(
            model_config.vaccine_catalog.get_product(
                model_config.resolved_product_id()
            ).dose_cost
        )
        dose_cost_results = {}
        for bound_name, bound_value in {
            "lower": dose_cost * 0.8,
            "upper": dose_cost * 1.2,
        }.items():
            perturbed_model_config = _replace_product_dose_cost(
                model_config,
                bound_value,
            )
            vaccinated_model_perturbed = _build_model(perturbed_model_config)
            vaccinated_simulation_perturbed = vaccinated_model_perturbed.simulate()
            dose_cost_results[bound_name] = _evaluate_from_existing_simulations(
                vaccinated_simulation_perturbed,
                reference_simulation,
                evaluation_config,
            )
        runtime_results["dose_cost"] = dose_cost_results

        discount_results = {}
        for bound_name, bound_value in {"lower": 0.01, "upper": 0.05}.items():
            perturbed_evaluation_config = _replace_evaluation_value(
                evaluation_config,
                "discount_rate",
                bound_value,
            )
            discount_results[bound_name] = _evaluate_from_existing_simulations(
                vaccinated_simulation,
                reference_simulation,
                perturbed_evaluation_config,
            )
        runtime_results["discount_rate"] = discount_results

        cost_per_cecx = float(evaluation_config.cost_per_cecx)
        cost_results = {}
        for bound_name, bound_value in {
            "lower": cost_per_cecx * 0.8,
            "upper": cost_per_cecx * 1.2,
        }.items():
            perturbed_evaluation_config = _replace_evaluation_value(
                evaluation_config,
                "cost_per_cecx",
                bound_value,
            )
            cost_results[bound_name] = _evaluate_from_existing_simulations(
                vaccinated_simulation,
                reference_simulation,
                perturbed_evaluation_config,
            )
        runtime_results["cost_per_cecx"] = cost_results

        daly_fatal = float(evaluation_config.daly_fatal)
        daly_results = {}
        for bound_name, bound_value in {
            "lower": daly_fatal * 0.9,
            "upper": daly_fatal * 1.1,
        }.items():
            perturbed_evaluation_config = _replace_evaluation_value(
                evaluation_config,
                "daly_fatal",
                bound_value,
            )
            daly_results[bound_name] = _evaluate_from_existing_simulations(
                vaccinated_simulation,
                reference_simulation,
                perturbed_evaluation_config,
            )
        runtime_results["daly_fatal"] = daly_results

        baseline_icur_value = baseline_icur
        for parameter in SENSITIVITY_PARAMETERS:
            original_value, lower_value, upper_value = _parameter_display_values(
                parameter,
                model_config,
                evaluation_config,
            )
            lower_icur_value = runtime_results[parameter.runtime_key]["lower"]
            upper_icur_value = runtime_results[parameter.runtime_key]["upper"]
            rows.append(
                {
                    "scenario_years": years,
                    "parameter": parameter.table_label,
                    "plot_label": parameter.plot_label,
                    "change_display": parameter.change_display,
                    "original_value_display": _format_parameter_value(
                        original_value,
                        parameter.value_unit,
                    ),
                    "lower_value_display": _format_parameter_value(
                        lower_value,
                        parameter.value_unit,
                    ),
                    "upper_value_display": _format_parameter_value(
                        upper_value,
                        parameter.value_unit,
                    ),
                    "baseline_icur": baseline_icur_value,
                    "lower_icur": lower_icur_value,
                    "upper_icur": upper_icur_value,
                    "absolute_change": max(
                        abs(lower_icur_value - baseline_icur_value),
                        abs(upper_icur_value - baseline_icur_value),
                    ),
                    "lower_delta": lower_icur_value - baseline_icur_value,
                    "upper_delta": upper_icur_value - baseline_icur_value,
                }
            )
    return scenarios, rows


def _sensitivity_payload_path(path: str | Path | None, output_dir: Path) -> Path:
    if path is not None:
        return Path(path)
    return output_dir / "sensitivity_s3.json"


def _write_sensitivity_outputs(
    scenarios: list[tuple[int, Path]],
    rows: list[dict[str, object]],
    output_dir: Path,
    *,
    payload_path: Path,
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "scenarios": [
            {"years": years, "directory": str(directory)}
            for years, directory in scenarios
        ],
        "rows": rows,
    }
    payload_path.parent.mkdir(parents=True, exist_ok=True)
    payload_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    longform_path = output_dir / "sensitivity_s3.csv"
    pd.DataFrame.from_records(rows).to_csv(longform_path, index=False)
    return payload_path, longform_path


def _load_sensitivity_payload(
    path: Path,
) -> tuple[list[tuple[int, Path]], list[dict[str, object]]]:
    if not path.exists():
        raise ValueError(
            f"sensitivity payload not found: {path}. "
            "Run `uv run summary.py sensitivity` first."
        )
    payload = json.loads(path.read_text(encoding="utf-8"))
    scenarios = [
        (int(item["years"]), Path(item["directory"])) for item in payload["scenarios"]
    ]
    rows = list(payload["rows"])
    return scenarios, rows


def _adjust_excel_widths_all_sheets(path: Path) -> None:
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            text_lengths = [
                len(str(cell.value)) for cell in column_cells if cell.value is not None
            ]
            width = max(text_lengths, default=0) + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width, 12), 36
            )
    workbook.save(path)


def _validate_budget_dir(search_dir: Path) -> None:
    required = (
        "best_model_config.json",
        "evaluation_config.json",
        "best_simulation.h5",
    )
    missing = [name for name in required if not (search_dir / name).exists()]
    if missing:
        raise ValueError(
            f"search result directory {search_dir} is missing required files: "
            f"{', '.join(missing)}"
        )


def _budget_workbook_path(output_dir: Path) -> Path:
    return output_dir / "budget_impact.xlsx"


def _copay_workbook_path(output_dir: Path) -> Path:
    return output_dir / "copay_impact.xlsx"


def _triparty_workbook_path(output_dir: Path) -> Path:
    return output_dir / "triparty_impact.xlsx"


def _interpolate_cumulative_to_year_grid(
    time: np.ndarray,
    values: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    horizon = int(round(float(time[-1])))
    year_grid = np.arange(0, horizon + 1, dtype=float)
    interpolated = np.vstack(
        [
            np.interp(year_grid, time, values[:, age_index])
            for age_index in range(values.shape[1])
        ]
    ).T
    return year_grid, interpolated


def _compute_budget_component_frame(
    sim_result: SimulationResult,
    evaluation_config: EvaluationConfig,
    *,
    vacc_insured: float = 1.0,
    vacc_reimburse: float = 1.0,
    minor_insured: float = 0.869,
    cecx_diag: float = 0.81,
    cecx_treatment: float = 0.9,
    cecx_insured: tuple[float, float] = (0.7569, 0.2431),
    cecx_reimburse: tuple[float, float] = (0.479, 0.6211),
    n_minor: int = 10,
) -> pd.DataFrame:
    model = sim_result.get_model()
    cumulative_cecx = np.asarray(
        model.cumulative_cecx(sim_result.cumulative),
        dtype=float,
    )
    cumulative_vaccinated = np.asarray(
        model.cumulative_vaccinated(sim_result.cumulative),
        dtype=float,
    )

    year_grid, cumulative_cecx_yearly = _interpolate_cumulative_to_year_grid(
        np.asarray(sim_result.time, dtype=float),
        cumulative_cecx,
    )
    _, cumulative_vaccinated_yearly = _interpolate_cumulative_to_year_grid(
        np.asarray(sim_result.time, dtype=float),
        cumulative_vaccinated,
    )

    annual_cecx = cumulative_cecx_yearly[1:] - cumulative_cecx_yearly[:-1]
    annual_vaccinated = (
        cumulative_vaccinated_yearly[1:] - cumulative_vaccinated_yearly[:-1]
    )
    annual_treated = annual_cecx * cecx_diag * cecx_treatment

    discount = np.power(
        1 / (1 + evaluation_config.discount_rate),
        np.arange(annual_cecx.shape[0]),
    )

    product_id = model.config.resolved_product_id()
    if product_id is None:
        vaccine_cost_per_age = np.zeros(model.nages, dtype=float)
    else:
        product = model.config.vaccine_catalog.get_product(product_id)
        vaccine_cost_per_age = model.vaccination_cost_per_age(
            dose_cost=product.dose_cost,
            dose_schedules=product.dose_schedules,
        )

    cost_vacc_age = (
        annual_vaccinated * vaccine_cost_per_age[None, :] * discount[:, None]
    )
    cost_cecx_age = annual_treated * evaluation_config.cost_per_cecx * discount[:, None]

    cost_vacc = cost_vacc_age.sum(axis=1)
    cost_cecx_minor = cost_cecx_age[:, :n_minor].sum(axis=1)
    cost_cecx_adult = cost_cecx_age[:, n_minor:].sum(axis=1)

    cost_vacc_insurance = cost_vacc * vacc_insured * vacc_reimburse
    cost_cecx_minor_insurance = cost_cecx_minor * minor_insured * cecx_reimburse[0]
    cost_cecx_adult_insurance = sum(
        cost_cecx_adult * insure_rate * reimburse_rate
        for insure_rate, reimburse_rate in zip(cecx_insured, cecx_reimburse)
    )
    cost_cecx_insured = cost_cecx_minor_insurance + cost_cecx_adult_insurance
    cost_insurance = cost_vacc_insurance + cost_cecx_insured

    years = year_grid[1:].astype(int) + BUDGET_BEGIN_YEAR - 1
    payload = {
        "Year": years,
        "Vaccination count": annual_vaccinated.sum(axis=1),
        "Treated cervical cancer cases": annual_treated.sum(axis=1),
        "Vaccination fund": cost_vacc_insurance / BUDGET_UNIT_DIVISOR,
        "Treatment fund": cost_cecx_insured / BUDGET_UNIT_DIVISOR,
        "Total fund": cost_insurance / BUDGET_UNIT_DIVISOR,
    }
    return pd.DataFrame(payload)


def _compute_triparty_component_frame(
    sim_result: SimulationResult,
    evaluation_config: EvaluationConfig,
    *,
    vacc_insured: float = 1.0,
    vacc_reimburse: float = TRIPARTY_MEDICAL_INSURANCE,
    gov_rate: float = TRIPARTY_GOVERNMENT,
    person_rate: float = TRIPARTY_INDIVIDUAL,
    minor_insured: float = 0.869,
    cecx_diag: float = 0.81,
    cecx_treatment: float = 0.9,
    cecx_insured: tuple[float, float] = (0.7569, 0.2431),
    cecx_reimburse: tuple[float, float] = (0.479, 0.6211),
    n_minor: int = 10,
) -> pd.DataFrame:
    model = sim_result.get_model()
    cumulative_cecx = np.asarray(
        model.cumulative_cecx(sim_result.cumulative),
        dtype=float,
    )
    cumulative_vaccinated = np.asarray(
        model.cumulative_vaccinated(sim_result.cumulative),
        dtype=float,
    )

    year_grid, cumulative_cecx_yearly = _interpolate_cumulative_to_year_grid(
        np.asarray(sim_result.time, dtype=float),
        cumulative_cecx,
    )
    _, cumulative_vaccinated_yearly = _interpolate_cumulative_to_year_grid(
        np.asarray(sim_result.time, dtype=float),
        cumulative_vaccinated,
    )

    annual_cecx = cumulative_cecx_yearly[1:] - cumulative_cecx_yearly[:-1]
    annual_vaccinated = (
        cumulative_vaccinated_yearly[1:] - cumulative_vaccinated_yearly[:-1]
    )
    annual_treated = annual_cecx * cecx_diag * cecx_treatment

    discount = np.power(
        1 / (1 + evaluation_config.discount_rate),
        np.arange(annual_cecx.shape[0]),
    )

    product_id = model.config.resolved_product_id()
    if product_id is None:
        vaccine_cost_per_age = np.zeros(model.nages, dtype=float)
    else:
        product = model.config.vaccine_catalog.get_product(product_id)
        vaccine_cost_per_age = model.vaccination_cost_per_age(
            dose_cost=product.dose_cost,
            dose_schedules=product.dose_schedules,
        )

    cost_vacc_age = (
        annual_vaccinated * vaccine_cost_per_age[None, :] * discount[:, None]
    )
    cost_cecx_age = annual_treated * evaluation_config.cost_per_cecx * discount[:, None]

    cost_vacc = cost_vacc_age.sum(axis=1)
    cost_cecx_minor = cost_cecx_age[:, :n_minor].sum(axis=1)
    cost_cecx_adult = cost_cecx_age[:, n_minor:].sum(axis=1)

    cost_vacc_insurance = cost_vacc * vacc_insured * vacc_reimburse
    cost_cecx_minor_insurance = cost_cecx_minor * minor_insured * cecx_reimburse[0]
    cost_cecx_adult_insurance = sum(
        cost_cecx_adult * insure_rate * reimburse_rate
        for insure_rate, reimburse_rate in zip(cecx_insured, cecx_reimburse)
    )
    cost_cecx_insured = cost_cecx_minor_insurance + cost_cecx_adult_insurance
    cost_insurance = cost_vacc_insurance + cost_cecx_insured

    years = year_grid[1:].astype(int) + BUDGET_BEGIN_YEAR - 1
    payload = {
        "Year": years,
        "Medical insurance - Vaccination cost": (
            cost_vacc_insurance / BUDGET_UNIT_DIVISOR
        ),
        "Medical insurance - Total expenditure": (cost_insurance / BUDGET_UNIT_DIVISOR),
        "Government": cost_vacc * gov_rate / BUDGET_UNIT_DIVISOR,
        "Individual": cost_vacc * person_rate / BUDGET_UNIT_DIVISOR,
    }
    return pd.DataFrame(payload)


def _build_budget_sheet_dataframe(search_dir: Path) -> tuple[int, pd.DataFrame]:
    _validate_budget_dir(search_dir)
    years = _scenario_years_from_name(search_dir)
    model_config = _load_model_config(search_dir / "best_model_config.json")
    evaluation_config = _load_evaluation_config(search_dir / "evaluation_config.json")

    vaccinated_simulation = SimulationResult.from_hdf(search_dir / "best_simulation.h5")
    reference_model = _build_model(_build_reference_config(model_config))
    reference_simulation = reference_model.simulate()

    vaccinated_df = _compute_budget_component_frame(
        vaccinated_simulation,
        evaluation_config,
    )
    reference_df = _compute_budget_component_frame(
        reference_simulation,
        evaluation_config,
    )

    merged = vaccinated_df.merge(
        reference_df,
        on="Year",
        how="inner",
        suffixes=(" - Optimal strategy", " - No vaccination"),
    )
    merged = merged.sort_values("Year").reset_index(drop=True)
    merged["Treatment fund increment"] = (
        merged["Treatment fund - Optimal strategy"]
        - merged["Treatment fund - No vaccination"]
    )
    merged["Total fund increment"] = (
        merged["Total fund - Optimal strategy"] - merged["Total fund - No vaccination"]
    )
    ordered = merged[
        [
            "Year",
            "Treatment fund - No vaccination",
            "Treatment fund - Optimal strategy",
            "Treatment fund increment",
            "Total fund - No vaccination",
            "Total fund - Optimal strategy",
            "Total fund increment",
            "Vaccination fund - Optimal strategy",
            "Vaccination count - Optimal strategy",
            "Treated cervical cancer cases - No vaccination",
            "Treated cervical cancer cases - Optimal strategy",
        ]
    ]
    return years, ordered


def _build_copay_sheet_dataframe(
    search_dir: Path,
    *,
    person_shares: tuple[float, float] | list[float],
) -> tuple[int, pd.DataFrame]:
    _validate_budget_dir(search_dir)
    years = _scenario_years_from_name(search_dir)
    model_config = _load_model_config(search_dir / "best_model_config.json")
    evaluation_config = _load_evaluation_config(search_dir / "evaluation_config.json")

    vaccinated_simulation = SimulationResult.from_hdf(search_dir / "best_simulation.h5")
    reference_model = _build_model(_build_reference_config(model_config))
    reference_simulation = reference_model.simulate()
    reference_df = _compute_budget_component_frame(
        reference_simulation,
        evaluation_config,
    )

    merged = reference_df.loc[:, ["Year", "Total fund"]].rename(
        columns={"Total fund": "No vaccination baseline fund"}
    )
    ordered_columns = ["Year"]
    for scheme_code, vacc_reimburse, _ in _build_copay_schemes(person_shares):
        scheme_df = _compute_budget_component_frame(
            vaccinated_simulation,
            evaluation_config,
            vacc_reimburse=vacc_reimburse,
        )
        scheme_df = scheme_df.loc[:, ["Year", "Vaccination fund", "Total fund"]].rename(
            columns={
                "Vaccination fund": f"{scheme_code} Vaccination cost",
                "Total fund": f"{scheme_code} Medical insurance fund expenditure",
            }
        )
        merged = merged.merge(scheme_df, on="Year", how="inner")
        merged[f"{scheme_code} Medical insurance increment"] = (
            merged[f"{scheme_code} Medical insurance fund expenditure"]
            - merged["No vaccination baseline fund"]
        )
        ordered_columns.extend(
            [
                f"{scheme_code} Vaccination cost",
                f"{scheme_code} Medical insurance fund expenditure",
                f"{scheme_code} Medical insurance increment",
            ]
        )

    ordered = merged[ordered_columns].copy()
    return years, ordered


def _build_triparty_sheet_dataframe(search_dir: Path) -> tuple[int, pd.DataFrame]:
    _validate_budget_dir(search_dir)
    years = _scenario_years_from_name(search_dir)
    model_config = _load_model_config(search_dir / "best_model_config.json")
    evaluation_config = _load_evaluation_config(search_dir / "evaluation_config.json")

    vaccinated_simulation = SimulationResult.from_hdf(search_dir / "best_simulation.h5")
    reference_model = _build_model(_build_reference_config(model_config))
    reference_simulation = reference_model.simulate()
    reference_df = _compute_budget_component_frame(
        reference_simulation,
        evaluation_config,
    )
    scheme_df = _compute_triparty_component_frame(
        vaccinated_simulation,
        evaluation_config,
    )
    merged = scheme_df.merge(
        reference_df.loc[:, ["Year", "Total fund"]].rename(
            columns={"Total fund": "No vaccination baseline fund"}
        ),
        on="Year",
        how="inner",
    )
    merged["Medical insurance - Increment"] = (
        merged["Medical insurance - Total expenditure"]
        - merged["No vaccination baseline fund"]
    )
    ordered = merged[
        [
            "Year",
            "Medical insurance - Vaccination cost",
            "Medical insurance - Total expenditure",
            "Medical insurance - Increment",
            "Government",
            "Individual",
        ]
    ].copy()
    total_row = {"Year": "Total"}
    for column in ordered.columns[1:]:
        total_row[column] = float(ordered[column].sum())
    ordered = pd.concat([ordered, pd.DataFrame([total_row])], ignore_index=True)
    return years, ordered


def _build_budget_readme_dataframe() -> pd.DataFrame:
    rows = [
        ("Workbook purpose", "Budget impact analysis for optimal strategies"),
        ("Currency output", "RMB"),
        ("Unit", BUDGET_UNIT_LABEL),
        ("Base year", str(BUDGET_BEGIN_YEAR)),
        ("vacc_insured", "1.0"),
        ("vacc_reimburse", "1.0"),
        ("minor_insured", "0.869"),
        ("cecx_diag", "0.81"),
        ("cecx_treatment", "0.9"),
        ("cecx_insured", "(0.7569, 0.2431)"),
        ("cecx_reimburse", "(0.479, 0.6211)"),
        ("discount_rate", "0.03"),
    ]
    return pd.DataFrame(rows, columns=["Item", "Value"])


def _build_copay_readme_dataframe(
    person_shares: tuple[float, float] | list[float],
) -> pd.DataFrame:
    schemes = _build_copay_schemes(person_shares)
    rows = [
        ("Workbook purpose", "Co-payment impact analysis for optimal strategies"),
        ("Currency output", "RMB"),
        ("Unit", BUDGET_UNIT_LABEL),
        ("Base year", str(BUDGET_BEGIN_YEAR)),
        *[
            (f"Scheme {scheme_code}", _copay_scheme_title(scheme_code))
            for scheme_code, _, _ in schemes
        ],
        (
            "Vaccination cost meaning",
            "Medical-insurance-paid vaccination expenditure under each scheme",
        ),
        (
            "Medical insurance fund expenditure meaning",
            "Total medical insurance fund expenditure with vaccination",
        ),
        (
            "Medical insurance increment meaning",
            "Total medical insurance fund expenditure increment vs no vaccination",
        ),
        ("vacc_insured", "1.0"),
        ("minor_insured", "0.869"),
        ("cecx_diag", "0.81"),
        ("cecx_treatment", "0.9"),
        ("cecx_insured", "(0.7569, 0.2431)"),
        ("cecx_reimburse", "(0.479, 0.6211)"),
        ("discount_rate", "0.03"),
    ]
    return pd.DataFrame(rows, columns=["Item", "Value"])


def _build_triparty_readme_dataframe() -> pd.DataFrame:
    rows = [
        (
            "Workbook purpose",
            "Multi-party funding analysis for optimal strategies",
        ),
        ("Currency output", "RMB"),
        ("Unit", BUDGET_UNIT_LABEL),
        ("Base year", str(BUDGET_BEGIN_YEAR)),
        (
            "Scheme",
            "medical insurance 27.91% / government 26.19% / individual 45.90%",
        ),
        ("vacc_insured", "1.0"),
        ("minor_insured", "0.869"),
        ("cecx_diag", "0.81"),
        ("cecx_treatment", "0.9"),
        ("cecx_insured", "(0.7569, 0.2431)"),
        ("cecx_reimburse", "(0.479, 0.6211)"),
        ("discount_rate", "0.03"),
        ("busi_rate", "0.0"),
    ]
    return pd.DataFrame(rows, columns=["Item", "Value"])


def _write_budget_workbook(
    output_dir: Path,
    sheet_payloads: list[tuple[int, pd.DataFrame]],
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = _budget_workbook_path(output_dir)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for years, dataframe in sheet_payloads:
            dataframe.to_excel(writer, sheet_name=f"{years}y", index=False)
        _build_budget_readme_dataframe().to_excel(
            writer,
            sheet_name="README",
            index=False,
        )
    _adjust_excel_widths_all_sheets(workbook_path)
    return workbook_path


def _write_copay_workbook(
    output_dir: Path,
    sheet_payloads: list[tuple[int, pd.DataFrame]],
    *,
    person_shares: tuple[float, float] | list[float],
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = _copay_workbook_path(output_dir)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for years, dataframe in sheet_payloads:
            dataframe.to_excel(writer, sheet_name=f"{years}y", index=False)
        _build_copay_readme_dataframe(person_shares).to_excel(
            writer,
            sheet_name="README",
            index=False,
        )
    _adjust_excel_widths_all_sheets(workbook_path)
    return workbook_path


def _write_triparty_workbook(
    output_dir: Path,
    sheet_payloads: list[tuple[int, pd.DataFrame]],
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = _triparty_workbook_path(output_dir)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for years, dataframe in sheet_payloads:
            dataframe.to_excel(writer, sheet_name=f"{years}y", index=False)
        _build_triparty_readme_dataframe().to_excel(
            writer,
            sheet_name="README",
            index=False,
        )
    _adjust_excel_widths_all_sheets(workbook_path)
    return workbook_path


def _load_budget_sheets(
    budget_path: Path,
) -> list[tuple[int, pd.DataFrame]]:
    if not budget_path.exists():
        raise ValueError(
            f"budget workbook not found: {budget_path}. "
            "Run `uv run summary.py budget` first."
        )
    workbook = pd.read_excel(budget_path, sheet_name=None)
    sheets: list[tuple[int, pd.DataFrame]] = []
    for name, dataframe in workbook.items():
        match = re.match(r"^(\d+)y$", name)
        if match is None:
            continue
        sheets.append((int(match.group(1)), dataframe))
    if not sheets:
        raise ValueError(f"no scenario sheets found in {budget_path}")
    return sorted(sheets, key=lambda item: item[0])


def _load_copay_sheets(
    copay_path: Path,
) -> list[tuple[int, pd.DataFrame]]:
    if not copay_path.exists():
        raise ValueError(
            f"co-payment workbook not found: {copay_path}. "
            "Run `uv run summary.py copay` first."
        )
    workbook = pd.read_excel(copay_path, sheet_name=None)
    sheets: list[tuple[int, pd.DataFrame]] = []
    for name, dataframe in workbook.items():
        match = re.match(r"^(\d+)y$", name)
        if match is None:
            continue
        sheets.append((int(match.group(1)), dataframe))
    if not sheets:
        raise ValueError(f"no scenario sheets found in {copay_path}")
    return sorted(sheets, key=lambda item: item[0])


def _load_triparty_sheets(
    triparty_path: Path,
) -> list[tuple[int, pd.DataFrame]]:
    if not triparty_path.exists():
        raise ValueError(
            f"multi-party workbook not found: {triparty_path}. "
            "Run `uv run summary.py triparty` first."
        )
    workbook = pd.read_excel(triparty_path, sheet_name=None)
    sheets: list[tuple[int, pd.DataFrame]] = []
    for name, dataframe in workbook.items():
        match = re.match(r"^(\d+)y$", name)
        if match is None:
            continue
        sheets.append((int(match.group(1)), dataframe))
    if not sheets:
        raise ValueError(f"no scenario sheets found in {triparty_path}")
    return sorted(sheets, key=lambda item: item[0])


def _build_tabs3_dataframe(
    scenarios: list[tuple[int, Path]],
    rows: list[dict[str, object]],
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for parameter in SENSITIVITY_PARAMETERS:
        matching_first = next(
            row
            for row in rows
            if row["scenario_years"] == scenarios[0][0]
            and row["parameter"] == parameter.table_label
        )
        record: dict[str, object] = {
            "Parameter": parameter.table_label,
            "The rate of change": matching_first["change_display"],
            "Original value": matching_first["original_value_display"],
            "Lower limit value": matching_first["lower_value_display"],
            "Upper limit value": matching_first["upper_value_display"],
        }
        for years, _ in scenarios:
            matching = next(
                row
                for row in rows
                if row["scenario_years"] == years
                and row["parameter"] == parameter.table_label
            )
            prefix = f"{years}y"
            record[f"{prefix} Original ICUR"] = _format_numeric(
                matching["baseline_icur"]
            )
            record[f"{prefix} Lower limit ICUR"] = _format_numeric(
                matching["lower_icur"]
            )
            record[f"{prefix} Upper limit ICUR"] = _format_numeric(
                matching["upper_icur"]
            )
            record[f"{prefix} Absolute change in ICUR"] = _format_numeric(
                matching["absolute_change"]
            )
        records.append(record)
    return pd.DataFrame.from_records(records)


def _load_search_results(
    results_root: Path, results_glob: str
) -> list[tuple[int, Path, SearchResult]]:
    search_dirs = _discover_search_dirs(results_root, results_glob)
    return [
        (
            _scenario_years_from_name(directory),
            directory,
            SearchResult.from_dir(directory),
        )
        for directory in search_dirs
    ]


def _subplot_panel_label(index: int) -> str:
    return chr(ord("A") + index)


def _format_heatmap_axis(
    ax: plt.Axes,
    *,
    time: np.ndarray,
    age_labels: list[str],
    show_xlabel: bool,
    show_ylabel: bool,
) -> None:
    ax.set_facecolor("white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.9)
    ax.spines["bottom"].set_linewidth(0.9)
    ax.tick_params(axis="both", labelsize=7.8, width=0.9, length=3, color="#444444")
    ax.grid(False)
    if show_xlabel:
        ax.set_xlabel("Time (years)", fontsize=9)
    if show_ylabel:
        tick_step = 2 if len(age_labels) <= 20 else 3
        tick_indices = list(range(0, len(age_labels), tick_step))
        if (len(age_labels) - 1) not in tick_indices:
            tick_indices.append(len(age_labels) - 1)
        ax.set_ylabel("Age group", fontsize=9)
        ax.set_yticks([index + 0.5 for index in tick_indices])
        ax.set_yticklabels([age_labels[index] for index in tick_indices])
    else:
        ax.set_yticks([])
    ax.set_xlim(float(time[0]), float(time[-1]))
    ax.set_ylim(0, len(age_labels))


def _build_fig2(
    output_dir: Path,
    scenarios: list[Fig2Scenario],
) -> Path:
    incidence_stack = np.concatenate(
        [scenario.incidence_reduction.ravel() for scenario in scenarios]
    )
    mortality_stack = np.concatenate(
        [scenario.mortality_reduction.ravel() for scenario in scenarios]
    )
    incidence_vmax = float(
        np.percentile(incidence_stack[np.isfinite(incidence_stack)], 99.5)
    )
    mortality_vmax = float(
        np.percentile(mortality_stack[np.isfinite(mortality_stack)], 99.5)
    )
    incidence_vmax = max(incidence_vmax, 1e-6)
    mortality_vmax = max(mortality_vmax, 1e-6)

    horizon_colors = {
        scenario.years: color
        for scenario, color in zip(
            scenarios,
            ["#2C7FB8", "#F28E2B", "#4E79A7", "#D1495B", "#7FB069", "#6A4C93"],
        )
    }

    fig = plt.figure(figsize=(16.6, 24.8))
    outer = fig.add_gridspec(
        5,
        1,
        height_ratios=[2.65, 2.65, 1.38, 1.28, 1.48],
        hspace=0.34,
    )
    grid_a = outer[0].subgridspec(2, 3, wspace=0.14, hspace=0.18)
    grid_b = outer[1].subgridspec(2, 3, wspace=0.14, hspace=0.18)
    grid_c = outer[2].subgridspec(1, 4, wspace=0.24)
    grid_d = outer[3].subgridspec(1, 3, wspace=0.24)
    grid_e = outer[4].subgridspec(1, 3, wspace=0.24)

    axes_a = np.array(
        [fig.add_subplot(grid_a[index // 3, index % 3]) for index in range(6)]
    )
    axes_b = np.array(
        [fig.add_subplot(grid_b[index // 3, index % 3]) for index in range(6)]
    )
    axes_c = np.array([fig.add_subplot(grid_c[0, index]) for index in range(4)])
    axes_d = np.array([fig.add_subplot(grid_d[0, index]) for index in range(3)])
    axes_e = np.array([fig.add_subplot(grid_e[0, index]) for index in range(3)])

    heatmap_a = None
    heatmap_b = None
    for index, (scenario, axis) in enumerate(zip(scenarios, axes_a)):
        heatmap_a = axis.imshow(
            scenario.incidence_reduction.T,
            origin="lower",
            aspect="auto",
            extent=(scenario.time[0], scenario.time[-1], 0, len(scenario.age_labels)),
            cmap="YlOrRd",
            vmin=0,
            vmax=incidence_vmax,
            interpolation="nearest",
        )
        axis.set_title(f"{scenario.years}-year horizon", fontsize=9.5, pad=4)
        if index == 0:
            axis.text(
                -0.24,
                1.12,
                "A",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )
        _format_heatmap_axis(
            axis,
            time=scenario.time,
            age_labels=scenario.age_labels,
            show_xlabel=index >= 3,
            show_ylabel=index % 3 == 0,
        )

    for index, (scenario, axis) in enumerate(zip(scenarios, axes_b)):
        heatmap_b = axis.imshow(
            scenario.mortality_reduction.T,
            origin="lower",
            aspect="auto",
            extent=(scenario.time[0], scenario.time[-1], 0, len(scenario.age_labels)),
            cmap="PuBuGn",
            vmin=0,
            vmax=mortality_vmax,
            interpolation="nearest",
        )
        axis.set_title(f"{scenario.years}-year horizon", fontsize=9.5, pad=4)
        if index == 0:
            axis.text(
                -0.24,
                1.12,
                "B",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )
        _format_heatmap_axis(
            axis,
            time=scenario.time,
            age_labels=scenario.age_labels,
            show_xlabel=index >= 3,
            show_ylabel=index % 3 == 0,
        )

    colorbar_a = fig.colorbar(
        heatmap_a,
        ax=axes_a.tolist(),
        fraction=0.018,
        pad=0.012,
    )
    colorbar_a.set_label(
        "Incidence reduction relative to no vaccination\n(/100,000 women)",
        fontsize=9,
    )
    colorbar_a.ax.tick_params(labelsize=8)

    colorbar_b = fig.colorbar(
        heatmap_b,
        ax=axes_b.tolist(),
        fraction=0.018,
        pad=0.012,
    )
    colorbar_b.set_label(
        "Mortality reduction relative to no vaccination\n(/100,000 women)",
        fontsize=9,
    )
    colorbar_b.ax.tick_params(labelsize=8)

    c_specs = (
        ("vaccine_number", "Vaccine number\n(10,000 doses)", 1.0),
        ("vaccine_cost", "Vaccine cost\n(10,000 yuan)", 1.0),
        ("cost_saved", "Cost saved\n(10,000 yuan)", 1.0),
        ("net_cost", "Net cost\n(10,000 yuan)", 1.0),
    )
    for index, (axis, (field, ylabel, divisor)) in enumerate(zip(axes_c, c_specs)):
        for scenario in scenarios:
            axis.plot(
                scenario.time,
                getattr(scenario, field) / divisor,
                color=horizon_colors[scenario.years],
                linewidth=1.8,
                label=f"{scenario.years} years",
            )
        axis.set_title(ylabel, fontsize=10, pad=4)
        axis.set_xlabel("Time (years)", fontsize=9.5)
        axis.set_ylabel("")
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "C",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )

    d_specs = (
        (
            "avoid_cecx",
            "Cases of cervical cancer prevented\n(10,000 cases)",
            10_000.0,
        ),
        (
            "avoid_cecx_deaths",
            "Cervical cancer deaths prevented\n(10,000 deaths)",
            10_000.0,
        ),
        ("avoid_daly", "Disability-adjusted life years saved", 1.0),
    )
    for index, (axis, (field, ylabel, divisor)) in enumerate(zip(axes_d, d_specs)):
        for scenario in scenarios:
            axis.plot(
                scenario.time,
                getattr(scenario, field) / divisor,
                color=horizon_colors[scenario.years],
                linewidth=1.8,
                label=f"{scenario.years} years",
            )
        axis.set_title(ylabel, fontsize=10, pad=4)
        axis.set_xlabel("Time (years)", fontsize=9.5)
        axis.set_ylabel("")
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "D",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )

    e_specs = (
        (
            "ic_per_cecx",
            "Incremental cost per case of cervical\ncancer prevention (yuan)",
            20.0,
        ),
        (
            "ic_per_cecx_death",
            "Incremental cost per cervical cancer\ndeath avoided (yuan)",
            30.0,
        ),
        (
            "icur",
            "Incremental cost per disability-adjusted\nlife year saved (yuan)",
            20.0,
        ),
    )
    for index, (axis, (field, ylabel, inset_start)) in enumerate(zip(axes_e, e_specs)):
        for scenario in scenarios:
            values = np.asarray(getattr(scenario, field), dtype=float)
            axis.plot(
                scenario.time,
                values,
                color=horizon_colors[scenario.years],
                linewidth=1.8,
                label=f"{scenario.years} years",
            )
        axis.set_title(ylabel, fontsize=10, pad=4)
        axis.set_xlabel("Time (years)", fontsize=9.5)
        axis.set_ylabel("")
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "E",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )

        inset = axis.inset_axes([0.18, 0.15, 0.72, 0.7])
        inset_values: list[np.ndarray] = []
        for scenario in scenarios:
            values = np.asarray(getattr(scenario, field), dtype=float)
            mask = np.isfinite(values) & (scenario.time >= inset_start)
            if not np.any(mask):
                continue
            inset.plot(
                scenario.time[mask],
                values[mask],
                color=horizon_colors[scenario.years],
                linewidth=1.2,
            )
            inset_values.append(values[mask])
        if inset_values:
            combined = np.concatenate(inset_values)
            finite = combined[np.isfinite(combined)]
            if finite.size > 0:
                lower = float(np.percentile(finite, 2))
                upper = float(np.percentile(finite, 98))
                if np.isclose(lower, upper):
                    margin = max(abs(lower) * 0.1, 1.0)
                    lower -= margin
                    upper += margin
                inset.set_xlim(
                    inset_start,
                    max(scenario.years for scenario in scenarios),
                )
                inset.set_ylim(lower, upper)
        inset.tick_params(labelsize=6.5, width=0.7, length=2.5)
        inset.spines["top"].set_visible(False)
        inset.spines["right"].set_visible(False)
        inset.grid(axis="y", color="#E1E6EF", linewidth=0.7, alpha=0.8)
        apply_scientific_format(inset, x=False, y=True)
        axis.indicate_inset_zoom(inset, edgecolor="#666666", alpha=0.7)

    legend_handles = [
        plt.Line2D(
            [0],
            [0],
            color=horizon_colors[scenario.years],
            linewidth=2.0,
            label=f"{scenario.years} years",
        )
        for scenario in scenarios
    ]
    fig.legend(
        handles=legend_handles,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.91),
        ncol=len(scenarios),
        frameon=False,
        fontsize=9,
    )
    a_title_y = max(axis.get_position().y1 for axis in axes_a[:3]) + 0.008
    b_title_y = max(axis.get_position().y1 for axis in axes_b[:3]) + 0.008
    fig.text(
        0.5,
        a_title_y,
        "Incidence reduction relative to no vaccination",
        ha="center",
        va="bottom",
        fontsize=11,
    )
    fig.text(
        0.5,
        b_title_y,
        "Mortality reduction relative to no vaccination",
        ha="center",
        va="bottom",
        fontsize=11,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_2.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _plot_price_demand_panel(
    ax: plt.Axes,
    curve_df: pd.DataFrame | None,
    method: str,
    selected_price: float | None,
    missing_message: str | None = None,
) -> None:
    colors = [
        "#3C3C3C",
        "#F28E2B",
        "#4E79A7",
        "#D1495B",
        "#7FB069",
        "#6A4C93",
        "#8C7A6B",
    ]

    if curve_df is None:
        ax.text(
            0.5,
            0.52,
            missing_message
            or (
                "price-demand curve data not available yet.\n"
                "Panel A will be rendered automatically\n"
                "once the required inputs are provided."
            ),
            transform=ax.transAxes,
            ha="center",
            va="center",
            fontsize=10,
            color="#555555",
        )
        ax.set_xlabel("Vaccine price (yuan)", fontsize=10)
        ax.set_ylabel("Vaccine demand", fontsize=10)
        ax.set_xlim(0, 3000)
        ax.set_ylim(0, 1.02)
        apply_nature_style(ax.figure, ax)
        return

    for income, color in zip(PRICE_DEMAND_INCOME_LEVELS, colors):
        income_curve = curve_df.loc[curve_df["income"] == income]
        if method == "empirical":
            ax.step(
                income_curve["price"].to_numpy(),
                income_curve["demand"].to_numpy(),
                where="post",
                color=color,
                linewidth=1.8,
                label=f"Income = {income:.0f}",
            )
        else:
            ax.plot(
                income_curve["price"].to_numpy(),
                income_curve["demand"].to_numpy(),
                color=color,
                linewidth=1.8,
                label=f"Income = {income:.0f}",
            )
    ax.set_xlabel("Vaccine price (yuan)", fontsize=10)
    ax.set_ylabel("Vaccine demand", fontsize=10)
    ax.set_xlim(0, float(curve_df["price"].max()))
    ax.set_ylim(0, 1.02)
    if selected_price is not None:
        ax.axvline(
            selected_price,
            color="black",
            linestyle="--",
            linewidth=1.3,
        )
    ax.legend(
        loc="upper right",
        frameon=False,
        fontsize=8.5,
        ncol=1,
    )
    apply_nature_style(ax.figure, ax)


def _price_demand_income_bin_edges() -> np.ndarray:
    midpoints = (PRICE_DEMAND_INCOME_LEVELS[:-1] + PRICE_DEMAND_INCOME_LEVELS[1:]) / 2
    return np.r_[-np.inf, midpoints, np.inf]


def _prepare_price_demand_source_data(
    price_demand_path: Path,
    search_dirs: list[Path],
    price_demand_column: str | None,
    price_demand_query: str | None,
    price_demand_doses: int | None,
    price_demand_wtp_scale: float,
) -> tuple[pd.DataFrame | None, dict[str, object] | None, str | None]:
    if not price_demand_path.exists():
        return (
            None,
            None,
            "price_demand_data.dta not available yet.\n"
            "Panel A will be rendered automatically\n"
            "once the data file is provided.",
        )

    product_id = _resolve_price_demand_product_id(search_dirs)
    raw_df = pd.read_stata(price_demand_path)
    if price_demand_query is not None:
        raw_df = raw_df.query(price_demand_query).copy()
    wtp_column = _resolve_price_demand_column(
        raw_df.columns,
        product_id,
        price_demand_column,
    )
    df = raw_df[[PRICE_DEMAND_INCOME_COLUMN, wtp_column]].rename(
        columns={PRICE_DEMAND_INCOME_COLUMN: "income", wtp_column: "WTP"}
    )
    df = df.query("income.notna() and WTP.notna() and WTP > 0").copy()
    plot_doses = _resolve_price_demand_plot_doses(search_dirs, price_demand_doses)
    # 问卷中的价格口径是三针总价，这里先换算成单针价格，
    # 再按指定剂次和行为折减系数映射回绘图口径。
    df["WTP"] = (df["WTP"] / 3.0) * float(plot_doses) * price_demand_wtp_scale

    selected_dose_price = _resolve_selected_dose_price(search_dirs)
    selected_price = selected_dose_price * float(plot_doses)
    max_observed_price = float(df["WTP"].max())
    plot_terminal_price = max(
        selected_price,
        max_observed_price + max(1.0, max_observed_price * 0.02),
    )
    metadata = {
        "product_id": product_id,
        "wtp_column": wtp_column,
        "query": price_demand_query,
        "plot_doses": plot_doses,
        "wtp_scale": price_demand_wtp_scale,
        "selected_dose_price": selected_dose_price,
        "selected_price": selected_price,
        "max_observed_price": max_observed_price,
        "plot_terminal_price": plot_terminal_price,
    }
    return df, metadata, None


def _empirical_price_demand_group_stats(
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[float, int]]:
    df = df.copy()
    df["income"] = pd.to_numeric(df["income"], errors="coerce")
    df["income_group"] = pd.cut(
        df["income"],
        bins=_price_demand_income_bin_edges(),
        labels=PRICE_DEMAND_INCOME_LEVELS,
        include_lowest=True,
    ).astype(float)
    group_sizes = (
        df.groupby("income_group", observed=False)
        .size()
        .reindex(PRICE_DEMAND_INCOME_LEVELS, fill_value=0)
    )
    return df, {float(index): int(value) for index, value in group_sizes.items()}


def _fit_weibull_price_demand(
    df: pd.DataFrame,
):
    try:
        from lifelines import WeibullAFTFitter
    except ImportError as exc:
        raise ImportError(
            "lifelines is required to fit the Weibull price-demand curve"
        ) from exc

    fit_df = df.copy()
    fit_df["status"] = 1
    threshold = np.quantile(fit_df["income"], 0.95)
    fit_df["income_wo_outlier"] = np.minimum(fit_df["income"], threshold)
    fitter = WeibullAFTFitter()
    fitter.fit(
        fit_df[["income_wo_outlier", "status", "WTP"]],
        duration_col="WTP",
        event_col="status",
    )
    return fitter, threshold


def _predict_weibull_price_demand(
    df: pd.DataFrame,
    *,
    income: float,
    price: float,
) -> dict[str, object]:
    fitter, threshold = _fit_weibull_price_demand(df)
    capped_income = min(float(income), float(threshold))
    demand = float(
        fitter.predict_survival_function(
            pd.DataFrame({"income_wo_outlier": [capped_income]}),
            times=[price],
        ).iloc[0, 0]
    )
    return {
        "income": float(income),
        "price": float(price),
        "demand": demand,
        "income_used_for_fit": capped_income,
    }


def _predict_weibull_price_from_demand(
    df: pd.DataFrame,
    *,
    income: float,
    demand: float,
) -> dict[str, object]:
    if not 0 < float(demand) <= 1:
        raise ValueError("target demand for Weibull inversion must be within (0, 1]")
    if float(demand) == 1.0:
        return {
            "income": float(income),
            "price": 0.0,
            "demand": float(demand),
            "income_used_for_fit": float(income),
        }

    from scipy.optimize import brentq

    fitter, threshold = _fit_weibull_price_demand(df)
    capped_income = min(float(income), float(threshold))

    def survival_at(price: float) -> float:
        return float(
            fitter.predict_survival_function(
                pd.DataFrame({"income_wo_outlier": [capped_income]}),
                times=[price],
            ).iloc[0, 0]
        )

    upper = max(float(df["WTP"].max()), 1.0)
    while survival_at(upper) > float(demand):
        upper *= 2.0

    price = float(brentq(lambda value: survival_at(value) - float(demand), 0.0, upper))
    return {
        "income": float(income),
        "price": price,
        "demand": float(demand),
        "income_used_for_fit": capped_income,
    }


def _predict_empirical_price_demand(
    df: pd.DataFrame,
    *,
    income: float,
    price: float,
) -> dict[str, object]:
    grouped_df, group_sizes = _empirical_price_demand_group_stats(df)
    income_group = pd.cut(
        pd.Series([income], dtype=float),
        bins=_price_demand_income_bin_edges(),
        labels=PRICE_DEMAND_INCOME_LEVELS,
        include_lowest=True,
    ).astype(float)
    representative_income = float(income_group.iloc[0])
    group_wtp = grouped_df.loc[
        grouped_df["income_group"] == representative_income,
        "WTP",
    ]
    if group_wtp.empty:
        raise ValueError(
            "no empirical price-demand samples found for income group "
            f"{representative_income}"
        )
    demand = float((group_wtp.to_numpy(dtype=float) >= price).mean())
    return {
        "income": float(income),
        "price": float(price),
        "demand": demand,
        "income_group": representative_income,
        "sample_size": group_sizes[representative_income],
    }


def _predict_empirical_price_from_demand(
    df: pd.DataFrame,
    *,
    income: float,
    demand: float,
) -> dict[str, object]:
    if not 0 < float(demand) <= 1:
        raise ValueError("target demand for empirical inversion must be within (0, 1]")

    grouped_df, group_sizes = _empirical_price_demand_group_stats(df)
    income_group = pd.cut(
        pd.Series([income], dtype=float),
        bins=_price_demand_income_bin_edges(),
        labels=PRICE_DEMAND_INCOME_LEVELS,
        include_lowest=True,
    ).astype(float)
    representative_income = float(income_group.iloc[0])
    group_wtp = grouped_df.loc[
        grouped_df["income_group"] == representative_income,
        "WTP",
    ]
    if group_wtp.empty:
        raise ValueError(
            "no empirical price-demand samples found for income group "
            f"{representative_income}"
        )

    sorted_wtp = np.sort(group_wtp.to_numpy(dtype=float))
    required_count = int(np.ceil(len(sorted_wtp) * float(demand)))
    price = float(sorted_wtp[len(sorted_wtp) - required_count])
    return {
        "income": float(income),
        "price": price,
        "demand": float(demand),
        "income_group": representative_income,
        "sample_size": group_sizes[representative_income],
        "price_definition": "maximum price with empirical demand >= target",
    }


def _prepare_weibull_price_demand_payloads(
    df: pd.DataFrame,
    search_dirs: list[Path],
    plot_terminal_price: float,
    plot_doses: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    fitter, _ = _fit_weibull_price_demand(df)

    price_grid = np.linspace(0.0, plot_terminal_price, 601)
    predictor = pd.DataFrame({"income_wo_outlier": PRICE_DEMAND_INCOME_LEVELS})
    survival = fitter.predict_survival_function(predictor, times=price_grid)
    survival.columns = PRICE_DEMAND_INCOME_LEVELS
    curve_df = (
        survival.rename_axis("price")
        .reset_index()
        .melt(id_vars="price", var_name="income", value_name="demand")
        .sort_values(["income", "price"], ignore_index=True)
    )
    curve_df["income"] = curve_df["income"].astype(float)
    curve_df["method"] = "weibull"

    point_rows: list[dict[str, object]] = []
    for search_dir in search_dirs:
        strategy = _load_search_strategy_metadata(search_dir)
        price = float(strategy["dose_price"]) * float(plot_doses)
        demand = fitter.predict_survival_function(
            predictor,
            times=[price],
        ).iloc[0]
        for income, income_demand in zip(
            PRICE_DEMAND_INCOME_LEVELS,
            demand,
            strict=True,
        ):
            point_rows.append(
                {
                    "Method": "weibull",
                    "Scenario": strategy["scenario"],
                    "Time horizon (years)": strategy["years"],
                    "Target age group": strategy["target_age_group"],
                    "Vaccine type": strategy["vaccine_type"],
                    "Plotted doses": plot_doses,
                    "Selected plotted price (yuan)": price,
                    "Household monthly income (yuan)": float(income),
                    "Demand": float(income_demand),
                }
            )
    point_df = pd.DataFrame.from_records(point_rows)
    return curve_df, point_df


def _prepare_empirical_price_demand_payloads(
    df: pd.DataFrame,
    search_dirs: list[Path],
    plot_terminal_price: float,
    plot_doses: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    grouped_df, group_sizes = _empirical_price_demand_group_stats(df)
    price_grid = np.unique(
        np.r_[0.0, grouped_df["WTP"].to_numpy(dtype=float), plot_terminal_price]
    )
    price_grid = price_grid[(price_grid >= 0.0) & (price_grid <= plot_terminal_price)]

    curve_rows: list[dict[str, object]] = []
    for income in PRICE_DEMAND_INCOME_LEVELS:
        income_group = grouped_df.loc[grouped_df["income_group"] == income, "WTP"]
        if income_group.empty:
            continue
        wtp_values = income_group.to_numpy(dtype=float)
        demand_values = np.array([(wtp_values >= price).mean() for price in price_grid])
        for price, demand in zip(price_grid, demand_values, strict=True):
            curve_rows.append(
                {
                    "method": "empirical",
                    "income": float(income),
                    "price": float(price),
                    "demand": float(demand),
                    "plotted_doses": plot_doses,
                    "sample_size": group_sizes[float(income)],
                }
            )
    curve_df = pd.DataFrame.from_records(curve_rows)

    point_rows: list[dict[str, object]] = []
    for search_dir in search_dirs:
        strategy = _load_search_strategy_metadata(search_dir)
        price = float(strategy["dose_price"]) * float(plot_doses)
        for income in PRICE_DEMAND_INCOME_LEVELS:
            income_group = grouped_df.loc[grouped_df["income_group"] == income, "WTP"]
            if income_group.empty:
                continue
            demand = float((income_group.to_numpy(dtype=float) >= price).mean())
            point_rows.append(
                {
                    "Method": "empirical",
                    "Scenario": strategy["scenario"],
                    "Time horizon (years)": strategy["years"],
                    "Target age group": strategy["target_age_group"],
                    "Vaccine type": strategy["vaccine_type"],
                    "Plotted doses": plot_doses,
                    "Selected plotted price (yuan)": price,
                    "Household monthly income (yuan)": float(income),
                    "Demand": demand,
                    "Sample size": group_sizes[float(income)],
                }
            )
    point_df = pd.DataFrame.from_records(point_rows)
    return curve_df, point_df


def _resolve_price_demand_product_id(search_dirs: list[Path]) -> str:
    product_ids = {
        str(_load_search_strategy_metadata(search_dir)["product_id"])
        for search_dir in search_dirs
    }
    if len(product_ids) != 1:
        raise ValueError(
            "price-demand analysis expects a single vaccine product across the "
            f"selected strategies, got: {sorted(product_ids)}"
        )
    return next(iter(product_ids))


def _resolve_selected_dose_price(search_dirs: list[Path]) -> float:
    prices = {
        float(_load_search_strategy_metadata(search_dir)["dose_price"])
        for search_dir in search_dirs
    }
    if len(prices) != 1:
        raise ValueError(
            "price-demand panel expects a single selected dose price across the "
            f"selected strategies, got: {sorted(prices)}"
        )
    return next(iter(prices))


def _resolve_selected_dose_count(search_dirs: list[Path]) -> int:
    dose_counts = {
        int(_load_search_strategy_metadata(search_dir)["dose_count"])
        for search_dir in search_dirs
    }
    if len(dose_counts) != 1:
        raise ValueError(
            "price-demand panel expects a single dose count across the selected "
            f"strategies, got: {sorted(dose_counts)}"
        )
    return next(iter(dose_counts))


def _resolve_price_demand_plot_doses(
    search_dirs: list[Path],
    explicit_doses: int | None,
) -> int:
    if explicit_doses is not None:
        return explicit_doses
    return _resolve_selected_dose_count(search_dirs)


def _resolve_price_demand_wtp_column(columns: pd.Index, product_id: str) -> str:
    candidates = PRICE_DEMAND_WTP_COLUMNS.get(product_id)
    if candidates is None:
        raise ValueError(
            f"no price-demand WTP column mapping configured for product {product_id!r}"
        )
    for candidate in candidates:
        if candidate in columns:
            return candidate
    raise ValueError(
        f"none of the candidate WTP columns for product {product_id!r} were found: "
        f"{list(candidates)}"
    )


def _resolve_price_demand_column(
    columns: pd.Index,
    product_id: str,
    explicit_column: str | None,
) -> str:
    if explicit_column is not None:
        if explicit_column not in columns:
            raise ValueError(
                f"requested price-demand column not found in Stata file: "
                f"{explicit_column!r}"
            )
        return explicit_column
    return _resolve_price_demand_wtp_column(columns, product_id)


def _prepare_price_demand_payloads(
    price_demand_path: Path,
    search_dirs: list[Path],
    method: str,
    price_demand_column: str | None,
    price_demand_query: str | None,
    price_demand_doses: int | None,
    price_demand_wtp_scale: float,
) -> tuple[pd.DataFrame | None, pd.DataFrame | None, str | None]:
    df, metadata, message = _prepare_price_demand_source_data(
        price_demand_path,
        search_dirs,
        price_demand_column,
        price_demand_query,
        price_demand_doses,
        price_demand_wtp_scale,
    )
    if df is None or metadata is None:
        return None, None, message

    try:
        if method == "empirical":
            curve_df, point_df = _prepare_empirical_price_demand_payloads(
                df,
                search_dirs,
                float(metadata["plot_terminal_price"]),
                int(metadata["plot_doses"]),
            )
        else:
            curve_df, point_df = _prepare_weibull_price_demand_payloads(
                df,
                search_dirs,
                float(metadata["plot_terminal_price"]),
                int(metadata["plot_doses"]),
            )
    except ImportError:
        return (
            None,
            None,
            "lifelines is required to fit the Weibull price-demand curve.\n"
            "Install it or switch to --price-demand-method empirical.",
        )

    curve_df["wtp_column"] = str(metadata["wtp_column"])
    curve_df["plotted_doses"] = int(metadata["plot_doses"])
    curve_df["query"] = "" if metadata["query"] is None else str(metadata["query"])
    curve_df["wtp_scale"] = float(metadata["wtp_scale"])
    point_df["WTP column"] = str(metadata["wtp_column"])
    point_df["Plotted doses"] = int(metadata["plot_doses"])
    point_df["Query"] = "" if metadata["query"] is None else str(metadata["query"])
    point_df["WTP scale"] = float(metadata["wtp_scale"])
    return curve_df, point_df, None


def _write_price_demand_outputs(
    output_dir: Path,
    curve_df: pd.DataFrame,
    point_df: pd.DataFrame,
    method: str,
    plot_doses: int,
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    curve_path = output_dir / f"price_demand_curve_{method}_{plot_doses}dose.csv"
    point_path = (
        output_dir / f"price_demand_optimal_strategies_{method}_{plot_doses}dose.csv"
    )
    curve_df.to_csv(curve_path, index=False)
    point_df.to_csv(point_path, index=False)
    return curve_path, point_path


def _build_fig3(
    output_dir: Path,
    budget_sheets: list[tuple[int, pd.DataFrame]],
    copay_sheets: list[tuple[int, pd.DataFrame]],
    triparty_sheets: list[tuple[int, pd.DataFrame]],
    price_demand_curve: pd.DataFrame | None,
    price_demand_method: str,
    plot_doses: int,
    selected_price: float | None,
    price_demand_message: str | None,
) -> Path:
    budget_by_year = dict(budget_sheets)
    copay_by_year = dict(copay_sheets)
    triparty_by_year = dict(triparty_sheets)
    horizons = sorted(set(budget_by_year) & set(copay_by_year) & set(triparty_by_year))
    if len(horizons) != 6:
        raise ValueError(
            "fig3 expects six shared horizons across budget, copay, and triparty "
            f"workbooks, got {horizons}"
        )

    fig = plt.figure(figsize=(16.4, 24.0))
    outer = fig.add_gridspec(
        4,
        1,
        height_ratios=[1.45, 2.25, 2.15, 2.15],
        hspace=0.32,
    )
    ax_a = fig.add_subplot(outer[0])
    grid_b = outer[1].subgridspec(2, 3, wspace=0.18, hspace=0.22)
    grid_c = outer[2].subgridspec(2, 3, wspace=0.18, hspace=0.22)
    grid_d = outer[3].subgridspec(2, 3, wspace=0.18, hspace=0.22)

    axes_b = np.array(
        [fig.add_subplot(grid_b[index // 3, index % 3]) for index in range(6)]
    )
    axes_c = np.array(
        [fig.add_subplot(grid_c[index // 3, index % 3]) for index in range(6)]
    )
    axes_d = np.array(
        [fig.add_subplot(grid_d[index // 3, index % 3]) for index in range(6)]
    )

    _plot_price_demand_panel(
        ax_a,
        price_demand_curve,
        price_demand_method,
        selected_price,
        price_demand_message,
    )
    ax_a.text(
        -0.05,
        1.04,
        "A",
        transform=ax_a.transAxes,
        fontsize=16,
        fontweight="bold",
        va="top",
        ha="left",
    )
    ax_a.set_title(
        "Vaccine demand curve across household income groups",
        fontsize=11,
        pad=4,
    )

    for index, (years, axis) in enumerate(zip(horizons, axes_b)):
        dataframe = budget_by_year[years]
        x = dataframe["Year"].to_numpy()
        axis.plot(
            x,
            dataframe["Treatment fund - No vaccination"].to_numpy(),
            color="#2C7FB8",
            linestyle="--",
            linewidth=1.4,
            label="Treatment fund - no vaccination",
        )
        axis.plot(
            x,
            dataframe["Treatment fund - Optimal strategy"].to_numpy(),
            color="#2C7FB8",
            linestyle="-",
            linewidth=1.8,
            label="Treatment fund - optimal strategy",
        )
        axis.plot(
            x,
            dataframe["Total fund - No vaccination"].to_numpy(),
            color="#D1495B",
            linestyle="--",
            linewidth=1.4,
            label="Total fund - no vaccination",
        )
        axis.plot(
            x,
            dataframe["Total fund - Optimal strategy"].to_numpy(),
            color="#D1495B",
            linestyle="-",
            linewidth=1.8,
            label="Total fund - optimal strategy",
        )
        axis.set_title(f"{years}-year horizon", fontsize=9.5, pad=4)
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "B",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )
            axis.legend(
                loc="upper right",
                frameon=False,
                fontsize=7.1,
                ncol=1,
            )
        if index >= 3:
            axis.set_xlabel("Calendar year", fontsize=9.5)

    for index, (years, axis) in enumerate(zip(horizons, axes_c)):
        dataframe = copay_by_year[years]
        scheme_meta = _copay_scheme_meta_from_dataframe(dataframe)
        if len(scheme_meta) != 2:
            raise ValueError(
                f"fig3 expects exactly two co-payment schemes, got {scheme_meta}"
            )
        x = dataframe["Year"].to_numpy()
        first_code, first_title = scheme_meta[0]
        second_code, second_title = scheme_meta[1]
        axis.plot(
            x,
            dataframe[f"{first_code} Medical insurance increment"].to_numpy(),
            color="#2C7FB8",
            linewidth=1.8,
            label=first_title,
        )
        axis.plot(
            x,
            dataframe[f"{second_code} Medical insurance increment"].to_numpy(),
            color="#F28E2B",
            linewidth=1.8,
            label=second_title,
        )
        axis.axhline(
            0.0,
            color="#7A7A7A",
            linewidth=1.0,
            linestyle="--",
            label="No additional spending",
        )
        axis.set_title(f"{years}-year horizon", fontsize=9.5, pad=4)
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "C",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )
            axis.legend(
                loc="upper right",
                frameon=False,
                fontsize=7.1,
                ncol=1,
            )
        if index >= 3:
            axis.set_xlabel("Calendar year", fontsize=9.5)

    for index, (years, axis) in enumerate(zip(horizons, axes_d)):
        dataframe = triparty_by_year[years]
        annual_df = dataframe.query("Year != 'Total'").copy()
        x = annual_df["Year"].astype(int).to_numpy()
        med = annual_df["Medical insurance - Vaccination cost"].to_numpy().cumsum()
        gov = annual_df["Government"].to_numpy().cumsum()
        person = annual_df["Individual"].to_numpy().cumsum()
        axis.plot(
            x,
            med,
            color="#4D4D4D",
            linewidth=1.8,
            label="Medical insurance",
        )
        axis.plot(
            x,
            gov,
            color="#F28E2B",
            linewidth=1.8,
            label="Government",
        )
        axis.plot(
            x,
            person,
            color="#4E79A7",
            linewidth=1.8,
            label="Individual",
        )
        axis.set_title(f"{years}-year horizon", fontsize=9.5, pad=4)
        apply_scientific_format(axis, x=False, y=True)
        apply_nature_style(fig, axis)
        if index == 0:
            axis.text(
                -0.22,
                1.12,
                "D",
                transform=axis.transAxes,
                fontsize=16,
                fontweight="bold",
                va="top",
                ha="left",
            )
            axis.legend(
                loc="upper left",
                frameon=False,
                fontsize=7.1,
                ncol=1,
            )
        if index >= 3:
            axis.set_xlabel("Calendar year", fontsize=9.5)

    for axis in np.r_[axes_b[:3], axes_c[:3], axes_d[:3]]:
        axis.tick_params(labelbottom=False)

    fig.text(
        0.5,
        max(axis.get_position().y1 for axis in axes_b[:3]) + 0.02,
        "Budget impact under optimal vaccination strategies",
        ha="center",
        va="bottom",
        fontsize=11,
    )
    fig.text(
        0.5,
        max(axis.get_position().y1 for axis in axes_c[:3]) + 0.02,
        "Incremental medical insurance expenditure under two co-payment schemes",
        ha="center",
        va="bottom",
        fontsize=11,
    )
    fig.text(
        0.5,
        max(axis.get_position().y1 for axis in axes_d[:3]) + 0.02,
        "Cumulative vaccine financing under the multi-party funding strategy",
        ha="center",
        va="bottom",
        fontsize=11,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_3.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    method_output_path = (
        output_dir / f"figure_3_{price_demand_method}_{plot_doses}dose.png"
    )
    if method_output_path != output_path:
        fig.savefig(method_output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return method_output_path


def _build_figs1(
    output_dir: Path, search_results: list[tuple[int, Path, SearchResult]]
) -> Path:
    fig, axes = plt.subplots(
        3,
        2,
        figsize=(8.4, 8.4),
        sharex=True,
        sharey=False,
    )
    axes_flat = axes.flatten()
    x_values: list[float] = []
    for _, _, result in search_results:
        completed_trials = [
            trial for trial in result.study.trials if trial.state.name == "COMPLETE"
        ]
        x_values.extend(
            result._trial_incidence(trial) * 1e5 for trial in completed_trials
        )
    x_min, x_max = min(x_values), max(x_values)
    x_pad = (x_max - x_min) * 0.05 if x_max > x_min else max(x_max * 0.05, 0.1)

    for index, ((years, _, result), axis) in enumerate(zip(search_results, axes_flat)):
        completed_trials = [
            trial for trial in result.study.trials if trial.state.name == "COMPLETE"
        ]
        product_ids = sorted(
            {
                str(trial.params.get("target_product_id", "unknown"))
                for trial in completed_trials
            }
        )
        pareto_numbers = {trial.number for trial in result.study.best_trials}
        for product_id in product_ids:
            product_trials = [
                trial
                for trial in completed_trials
                if trial.params.get("target_product_id") == product_id
            ]
            if not product_trials:
                continue
            axis.scatter(
                [result._trial_incidence(trial) * 1e5 for trial in product_trials],
                [result._trial_icur(trial) for trial in product_trials],
                color=PRODUCT_COLORS.get(product_id, "#6C757D"),
                alpha=0.7,
                s=16,
                linewidths=0,
            )
            highlight_trials = [
                trial for trial in product_trials if trial.number in pareto_numbers
            ]
            if highlight_trials:
                axis.scatter(
                    [
                        result._trial_incidence(trial) * 1e5
                        for trial in highlight_trials
                    ],
                    [result._trial_icur(trial) for trial in highlight_trials],
                    color=PRODUCT_COLORS.get(product_id, "#6C757D"),
                    marker="*",
                    s=56,
                    edgecolors="#202020",
                    linewidths=0.4,
                )
                axis.axvline(
                    x=4, color="#202020", linestyle="--", linewidth=0.8, alpha=0.9
                )
        axis.set_title(f"{years}-year horizon", fontsize=10, pad=4)
        axis.text(
            0.02,
            0.98,
            _subplot_panel_label(index),
            transform=axis.transAxes,
            va="top",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )
        axis.set_xlim(x_min - x_pad, x_max + x_pad)
        # axis.set_ylim(max(0, y_min - y_pad), y_max + y_pad)
        apply_scientific_format(axis, x=False, y=True)
    for axis in axes_flat[:4]:
        axis.tick_params(labelbottom=False)
    fig.supxlabel("Final cervical cancer incidence (/100,000 women)", fontsize=11)
    fig.supylabel("ICUR", fontsize=11)
    all_product_ids = sorted(
        {
            str(trial.params.get("target_product_id", "unknown"))
            for _, _, result in search_results
            for trial in result.study.trials
            if trial.state.name == "COMPLETE"
        }
    )
    handles = [
        plt.Line2D(
            [0],
            [0],
            marker="o",
            linestyle="",
            markerfacecolor=PRODUCT_COLORS.get(product_id, "#6C757D"),
            markeredgecolor="none",
            markersize=6,
            label=_product_label(product_id),
        )
        for product_id in all_product_ids
    ]
    handles.extend(
        [
            plt.Line2D(
                [0],
                [0],
                marker="o",
                linestyle="",
                markerfacecolor="#444444",
                markeredgecolor="none",
                markersize=6,
                label="All trials",
            ),
            plt.Line2D(
                [0],
                [0],
                marker="*",
                linestyle="",
                markerfacecolor="#444444",
                markeredgecolor="#202020",
                markersize=9,
                label="Pareto front",
            ),
        ]
    )
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=5,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    apply_nature_style(fig, axes_flat)
    fig.tight_layout(rect=(0, 0, 1, 0.95))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s1.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _build_figs2(
    output_dir: Path, search_results: list[tuple[int, Path, SearchResult]]
) -> Path:
    fig, axes = plt.subplots(3, 2, figsize=(8.6, 8.6), sharex=False)
    axes_flat = axes.flatten()
    right_axes = []
    for index, ((years, _, result), axis) in enumerate(zip(search_results, axes_flat)):
        completed_trials = [
            trial for trial in result.study.trials if trial.state.name == "COMPLETE"
        ]
        trial_numbers = [trial.number for trial in completed_trials]
        icur_values = [result._trial_icur(trial) for trial in completed_trials]
        incidence_values = [
            result._trial_incidence(trial) * 1e5 for trial in completed_trials
        ]
        right_axis = axis.twinx()
        right_axes.append(right_axis)
        axis.plot(
            trial_numbers,
            icur_values,
            color="#2C7FB8",
            linewidth=1.1,
            alpha=0.95,
        )
        right_axis.plot(
            trial_numbers,
            incidence_values,
            color="#D1495B",
            linewidth=1.1,
            alpha=0.90,
        )
        axis.set_title(f"{years}-year horizon", fontsize=10, pad=4)
        axis.text(
            0.02,
            0.98,
            _subplot_panel_label(index),
            transform=axis.transAxes,
            va="top",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )
        axis.tick_params(axis="y", colors="#2C7FB8")
        right_axis.tick_params(axis="y", colors="#D1495B")
        apply_scientific_format(axis, x=False, y=True)
        apply_scientific_format(right_axis, x=False, y=True)
        apply_nature_style(fig, axis)
        right_axis.spines["top"].set_visible(False)
        right_axis.spines["left"].set_visible(False)
        right_axis.spines["right"].set_linewidth(1.0)
        right_axis.grid(False)
        if index % 2 == 0:
            right_axis.tick_params(labelright=False)
    for axis in axes_flat[:4]:
        axis.tick_params(labelbottom=False)
    for axis in axes_flat[::2]:
        axis.set_ylabel("ICUR", color="#2C7FB8", fontsize=9)
    for right_axis in right_axes[1::2]:
        right_axis.set_ylabel(
            "Final incidence (/100,000 women)", color="#D1495B", fontsize=9
        )
    fig.supxlabel("Trial", fontsize=11)
    handles = [
        plt.Line2D([0], [0], color="#2C7FB8", linewidth=1.6, label="ICUR"),
        plt.Line2D(
            [0],
            [0],
            color="#D1495B",
            linewidth=1.6,
            label="Final cervical cancer incidence (/100,000 women)",
        ),
    ]
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=2,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    fig.tight_layout(rect=(0, 0, 1, 0.96))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s2.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _build_figs3(
    output_dir: Path,
    scenarios: list[tuple[int, Path]],
    rows: list[dict[str, object]],
) -> Path:
    fig, axes = plt.subplots(3, 2, figsize=(13.2, 10.6), sharex=False, sharey=False)
    axes_flat = axes.flatten()
    lower_color = "#2C7FB8"
    upper_color = "#D1495B"

    for index, ((years, _), axis) in enumerate(zip(scenarios, axes_flat)):
        scenario_rows = [row for row in rows if row["scenario_years"] == years]
        scenario_rows = sorted(
            scenario_rows,
            key=lambda row: float(row["absolute_change"]),
            reverse=True,
        )
        y_positions = list(range(len(scenario_rows)))
        axis.barh(
            [y - 0.16 for y in y_positions],
            [row["lower_delta"] for row in scenario_rows],
            height=0.28,
            color=lower_color,
            alpha=0.85,
        )
        axis.barh(
            [y + 0.16 for y in y_positions],
            [row["upper_delta"] for row in scenario_rows],
            height=0.28,
            color=upper_color,
            alpha=0.85,
        )
        axis.axvline(0, color="#202020", linewidth=0.9, linestyle="--", alpha=0.9)
        axis.set_yticks(y_positions)
        axis.set_yticklabels(
            [row["plot_label"] for row in scenario_rows],
            fontsize=8.4,
            linespacing=1.05,
        )
        axis.tick_params(axis="y", pad=3)
        axis.invert_yaxis()
        axis.set_title(f"{years}-year horizon", fontsize=10, pad=4)
        axis.text(
            0.02,
            0.98,
            _subplot_panel_label(index),
            transform=axis.transAxes,
            va="top",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )
        apply_scientific_format(axis, x=True, y=False)

    fig.supxlabel("ΔICUR (yuan / DALY)", fontsize=11)
    handles = [
        plt.Line2D([0], [0], color=lower_color, linewidth=6, label="Lower limit"),
        plt.Line2D([0], [0], color=upper_color, linewidth=6, label="Upper limit"),
    ]
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=2,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    apply_nature_style(fig, axes_flat)
    fig.tight_layout(rect=(0.09, 0, 1, 0.96))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s3.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _build_figs4(
    output_dir: Path,
    budget_sheets: list[tuple[int, pd.DataFrame]],
) -> Path:
    fig, axes = plt.subplots(6, 2, figsize=(11.8, 15.2), sharex=False, sharey=False)
    treatment_color = "#2C7FB8"
    total_color = "#D1495B"
    axes_array = np.asarray(axes)

    for index, (years, dataframe) in enumerate(budget_sheets):
        treatment_axis = axes_array[index, 0]
        total_axis = axes_array[index, 1]

        treatment_axis.plot(
            dataframe["Year"],
            dataframe["Treatment fund - No vaccination"],
            color=treatment_color,
            linestyle="--",
            linewidth=1.8,
        )
        treatment_axis.plot(
            dataframe["Year"],
            dataframe["Treatment fund - Optimal strategy"],
            color=treatment_color,
            linestyle="-",
            linewidth=1.8,
        )
        total_axis.plot(
            dataframe["Year"],
            dataframe["Total fund - No vaccination"],
            color=total_color,
            linestyle="--",
            linewidth=1.8,
        )
        total_axis.plot(
            dataframe["Year"],
            dataframe["Total fund - Optimal strategy"],
            color=total_color,
            linestyle="-",
            linewidth=1.8,
        )

        treatment_axis.set_title(
            f"{years}-year horizon: Treatment fund",
            fontsize=9.5,
            pad=4,
        )
        total_axis.set_title(f"{years}-year horizon: Total fund", fontsize=9.5, pad=4)
        treatment_axis.text(
            0.02,
            0.98,
            _subplot_panel_label(index * 2),
            transform=treatment_axis.transAxes,
            va="top",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )
        total_axis.text(
            0.02,
            0.98,
            _subplot_panel_label(index * 2 + 1),
            transform=total_axis.transAxes,
            va="top",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )
        treatment_axis.set_ylabel(f"Expenditure ({BUDGET_UNIT_LABEL})", fontsize=9)
        total_axis.set_ylabel(f"Expenditure ({BUDGET_UNIT_LABEL})", fontsize=9)
        if index == len(budget_sheets) - 1:
            treatment_axis.set_xlabel("Calendar year", fontsize=9)
            total_axis.set_xlabel("Calendar year", fontsize=9)

        apply_scientific_format(treatment_axis, x=False, y=True)
        apply_scientific_format(total_axis, x=False, y=True)

    handles = [
        plt.Line2D(
            [0],
            [0],
            color="#444444",
            linestyle="--",
            linewidth=1.8,
            label="No vaccination",
        ),
        plt.Line2D(
            [0],
            [0],
            color="#444444",
            linestyle="-",
            linewidth=1.8,
            label="Optimal strategy",
        ),
    ]
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=2,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    apply_nature_style(fig, axes_array.flatten())
    fig.tight_layout(rect=(0, 0, 1, 0.975))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s4.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _build_figs5(
    output_dir: Path,
    copay_sheets: list[tuple[int, pd.DataFrame]],
) -> Path:
    fig, axes = plt.subplots(6, 2, figsize=(12.4, 15.2), sharex=False, sharey=False)
    axes_array = np.asarray(axes)
    vacc_color = "#2C7FB8"
    fund_color = "#D1495B"
    increment_color = "#2F855A"
    scheme_meta = _copay_scheme_meta_from_dataframe(copay_sheets[0][1])

    for row_index, (years, dataframe) in enumerate(copay_sheets):
        for col_index, (scheme_code, scheme_title) in enumerate(scheme_meta):
            axis = axes_array[row_index, col_index]
            axis.plot(
                dataframe["Year"],
                dataframe[f"{scheme_code} Vaccination cost"],
                color=vacc_color,
                linestyle="-",
                linewidth=1.6,
            )
            axis.plot(
                dataframe["Year"],
                dataframe[f"{scheme_code} Medical insurance fund expenditure"],
                color=fund_color,
                linestyle="-",
                linewidth=1.6,
            )
            axis.plot(
                dataframe["Year"],
                dataframe[f"{scheme_code} Medical insurance increment"],
                color=increment_color,
                linestyle="-",
                linewidth=1.6,
            )
            axis.set_title(
                f"{years}-year horizon: {scheme_title}",
                fontsize=9.2,
                pad=4,
            )
            axis.text(
                0.02,
                0.98,
                _subplot_panel_label(row_index * 2 + col_index),
                transform=axis.transAxes,
                va="top",
                ha="left",
                fontsize=10,
                fontweight="bold",
            )
            if col_index == 0:
                axis.set_ylabel(f"Expenditure ({BUDGET_UNIT_LABEL})", fontsize=9)
            if row_index == len(copay_sheets) - 1:
                axis.set_xlabel("Calendar year", fontsize=9)
            apply_scientific_format(axis, x=False, y=True)

    handles = [
        plt.Line2D(
            [0],
            [0],
            color=vacc_color,
            linestyle="-",
            linewidth=1.8,
            label="Vaccination cost",
        ),
        plt.Line2D(
            [0],
            [0],
            color=fund_color,
            linestyle="-",
            linewidth=1.8,
            label="Medical insurance fund expenditure",
        ),
        plt.Line2D(
            [0],
            [0],
            color=increment_color,
            linestyle="-",
            linewidth=1.8,
            label="Medical insurance increment",
        ),
    ]
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    apply_nature_style(fig, axes_array.flatten())
    fig.tight_layout(rect=(0, 0, 1, 0.975))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s5.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _build_figs6(
    output_dir: Path,
    triparty_sheets: list[tuple[int, pd.DataFrame]],
) -> Path:
    fig, axes = plt.subplots(6, 3, figsize=(14.2, 15.8), sharex=False, sharey=False)
    axes_array = np.asarray(axes)
    vacc_color = "#2C7FB8"
    total_color = "#D1495B"
    increment_color = "#2F855A"
    government_color = "#B7791F"
    individual_color = "#6B46C1"

    for index, (years, dataframe) in enumerate(triparty_sheets):
        annual_df = dataframe[
            pd.to_numeric(dataframe["Year"], errors="coerce").notna()
        ].copy()
        annual_df["Year"] = annual_df["Year"].astype(int)
        medical_axis = axes_array[index, 0]
        government_axis = axes_array[index, 1]
        individual_axis = axes_array[index, 2]

        medical_axis.plot(
            annual_df["Year"],
            annual_df["Medical insurance - Vaccination cost"],
            color=vacc_color,
            linestyle="-",
            linewidth=1.5,
        )
        medical_axis.plot(
            annual_df["Year"],
            annual_df["Medical insurance - Total expenditure"],
            color=total_color,
            linestyle="-",
            linewidth=1.5,
        )
        medical_axis.plot(
            annual_df["Year"],
            annual_df["Medical insurance - Increment"],
            color=increment_color,
            linestyle="-",
            linewidth=1.5,
        )
        government_axis.plot(
            annual_df["Year"],
            annual_df["Government"],
            color=government_color,
            linestyle="-",
            linewidth=1.5,
        )
        individual_axis.plot(
            annual_df["Year"],
            annual_df["Individual"],
            color=individual_color,
            linestyle="-",
            linewidth=1.5,
        )

        for col_index, (axis, title_suffix) in enumerate(
            (
                (medical_axis, "Medical insurance"),
                (government_axis, "Government"),
                (individual_axis, "Individual"),
            )
        ):
            axis.set_title(f"{years}-year horizon: {title_suffix}", fontsize=9.2, pad=4)
            axis.text(
                0.02,
                0.98,
                _subplot_panel_label(index * 3 + col_index),
                transform=axis.transAxes,
                va="top",
                ha="left",
                fontsize=10,
                fontweight="bold",
            )
            if col_index == 0:
                axis.set_ylabel(f"Expenditure ({BUDGET_UNIT_LABEL})", fontsize=9)
            if index == len(triparty_sheets) - 1:
                axis.set_xlabel("Calendar year", fontsize=9)
            apply_scientific_format(axis, x=False, y=True)

    handles = [
        plt.Line2D(
            [0],
            [0],
            color=vacc_color,
            linestyle="-",
            linewidth=1.8,
            label="Medical insurance - Vaccination cost",
        ),
        plt.Line2D(
            [0],
            [0],
            color=total_color,
            linestyle="-",
            linewidth=1.8,
            label="Medical insurance - Total expenditure",
        ),
        plt.Line2D(
            [0],
            [0],
            color=increment_color,
            linestyle="-",
            linewidth=1.8,
            label="Medical insurance - Increment",
        ),
        plt.Line2D(
            [0],
            [0],
            color=government_color,
            linestyle="-",
            linewidth=1.8,
            label="Government",
        ),
        plt.Line2D(
            [0],
            [0],
            color=individual_color,
            linestyle="-",
            linewidth=1.8,
            label="Individual",
        ),
    ]
    fig.legend(
        handles=handles,
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.995),
        fontsize=8.5,
    )
    apply_nature_style(fig, axes_array.flatten())
    fig.tight_layout(rect=(0, 0, 1, 0.975))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "figure_s6.png"
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _run_tab1(args: argparse.Namespace) -> None:
    results_root = Path(args.results_root)
    output_dir = Path(args.output_dir)
    dataframe = _build_tab1_dataframe(results_root, args.results_glob)
    excel_path, csv_path, markdown_path = _write_table_outputs(
        dataframe, output_dir, "table1"
    )
    print(f"Wrote {excel_path}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {markdown_path}")


def _run_tabs1(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    model_config = _load_model_config(Path(args.model_config))
    dataframe = _build_tabs1_dataframe(model_config)
    excel_path, csv_path, markdown_path = _write_table_outputs(
        dataframe, output_dir, "table_s1"
    )
    print(f"Wrote {excel_path}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {markdown_path}")


def _run_fig2(args: argparse.Namespace) -> None:
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    scenarios = [_build_fig2_scenario(search_dir) for search_dir in search_dirs]
    output_path = _build_fig2(Path(args.output_dir), scenarios)
    print(f"Wrote {output_path}")


def _run_fig3(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    plotted_doses = _resolve_price_demand_plot_doses(
        search_dirs,
        args.price_demand_doses,
    )
    selected_price = _resolve_selected_dose_price(search_dirs) * float(plotted_doses)
    budget_sheets = _load_budget_sheets(Path(args.budget_path))
    copay_sheets = _load_copay_sheets(Path(args.copay_path))
    triparty_sheets = _load_triparty_sheets(Path(args.triparty_path))
    price_demand_curve, price_demand_points, price_demand_message = (
        _prepare_price_demand_payloads(
            Path(args.price_demand_path),
            search_dirs,
            args.price_demand_method,
            args.price_demand_column,
            args.price_demand_query,
            args.price_demand_doses,
            args.price_demand_wtp_scale,
        )
    )
    if price_demand_curve is not None and price_demand_points is not None:
        curve_path, point_path = _write_price_demand_outputs(
            output_dir,
            price_demand_curve,
            price_demand_points,
            args.price_demand_method,
            plotted_doses,
        )
        print(f"Wrote {curve_path}")
        print(f"Wrote {point_path}")
    output_path = _build_fig3(
        output_dir,
        budget_sheets,
        copay_sheets,
        triparty_sheets,
        price_demand_curve,
        args.price_demand_method,
        plotted_doses,
        selected_price,
        price_demand_message,
    )
    print(f"Wrote {output_path}")


def _run_price_demand(args: argparse.Namespace) -> None:
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    df, metadata, message = _prepare_price_demand_source_data(
        Path(args.price_demand_path),
        search_dirs,
        args.price_demand_column,
        args.price_demand_query,
        args.price_demand_doses,
        args.price_demand_wtp_scale,
    )
    if df is None or metadata is None:
        raise ValueError(message or "price-demand data could not be prepared")

    if args.price is not None:
        if args.price_demand_method == "empirical":
            prediction = _predict_empirical_price_demand(
                df,
                income=float(args.income),
                price=float(args.price),
            )
        else:
            prediction = _predict_weibull_price_demand(
                df,
                income=float(args.income),
                price=float(args.price),
            )
        payload = {
            "method": args.price_demand_method,
            "income": float(args.income),
            "price": float(args.price),
            "demand": float(prediction["demand"]),
            "wtp_column": str(metadata["wtp_column"]),
            "query": args.price_demand_query,
            "plotted_doses": int(metadata["plot_doses"]),
            "wtp_scale": float(metadata["wtp_scale"]),
            "selected_price": float(metadata["selected_price"]),
        }
    else:
        if args.price_demand_method == "empirical":
            prediction = _predict_empirical_price_from_demand(
                df,
                income=float(args.income),
                demand=float(args.demand),
            )
        else:
            prediction = _predict_weibull_price_from_demand(
                df,
                income=float(args.income),
                demand=float(args.demand),
            )
        payload = {
            "method": args.price_demand_method,
            "income": float(args.income),
            "target_demand": float(args.demand),
            "price": float(prediction["price"]),
            "wtp_column": str(metadata["wtp_column"]),
            "query": args.price_demand_query,
            "plotted_doses": int(metadata["plot_doses"]),
            "wtp_scale": float(metadata["wtp_scale"]),
            "selected_price": float(metadata["selected_price"]),
        }
    for key in (
        "income_group",
        "sample_size",
        "income_used_for_fit",
        "price_definition",
    ):
        if key in prediction:
            payload[key] = prediction[key]
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def _run_figs1(args: argparse.Namespace) -> None:
    search_results = _load_search_results(Path(args.results_root), args.results_glob)
    output_path = _build_figs1(Path(args.output_dir), search_results)
    print(f"Wrote {output_path}")


def _run_figs2(args: argparse.Namespace) -> None:
    search_results = _load_search_results(Path(args.results_root), args.results_glob)
    output_path = _build_figs2(Path(args.output_dir), search_results)
    print(f"Wrote {output_path}")


def _run_sensitivity(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    scenarios, rows = _compute_sensitivity_payloads(
        Path(args.results_root),
        args.results_glob,
    )
    payload_path = _sensitivity_payload_path(None, output_dir)
    json_path, csv_path = _write_sensitivity_outputs(
        scenarios,
        rows,
        output_dir,
        payload_path=payload_path,
    )
    print(f"Wrote {json_path}")
    print(f"Wrote {csv_path}")


def _run_tabs3(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    payload_path = _sensitivity_payload_path(args.sensitivity_path, output_dir)
    scenarios, rows = _load_sensitivity_payload(payload_path)
    dataframe = _build_tabs3_dataframe(scenarios, rows)
    excel_path, csv_path, markdown_path = _write_table_outputs(
        dataframe, output_dir, "table_s3"
    )
    print(f"Wrote {excel_path}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {markdown_path}")


def _run_figs3(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    payload_path = _sensitivity_payload_path(args.sensitivity_path, output_dir)
    scenarios, rows = _load_sensitivity_payload(payload_path)
    output_path = _build_figs3(
        output_dir,
        scenarios,
        rows,
    )
    print(f"Wrote {output_path}")


def _run_budget(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    sheet_payloads = [
        _build_budget_sheet_dataframe(search_dir) for search_dir in search_dirs
    ]
    workbook_path = _write_budget_workbook(output_dir, sheet_payloads)
    print(f"Wrote {workbook_path}")


def _run_figs4(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    budget_sheets = _load_budget_sheets(Path(args.budget_path))
    output_path = _build_figs4(output_dir, budget_sheets)
    print(f"Wrote {output_path}")


def _run_copay(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    sheet_payloads = [
        _build_copay_sheet_dataframe(
            search_dir,
            person_shares=tuple(args.copay_person_shares),
        )
        for search_dir in search_dirs
    ]
    workbook_path = _write_copay_workbook(
        output_dir,
        sheet_payloads,
        person_shares=tuple(args.copay_person_shares),
    )
    print(f"Wrote {workbook_path}")


def _run_figs5(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    copay_sheets = _load_copay_sheets(Path(args.copay_path))
    output_path = _build_figs5(output_dir, copay_sheets)
    print(f"Wrote {output_path}")


def _run_triparty(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    search_dirs = _discover_search_dirs(Path(args.results_root), args.results_glob)
    sheet_payloads = [
        _build_triparty_sheet_dataframe(search_dir) for search_dir in search_dirs
    ]
    workbook_path = _write_triparty_workbook(output_dir, sheet_payloads)
    print(f"Wrote {workbook_path}")


def _run_figs6(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    triparty_sheets = _load_triparty_sheets(Path(args.triparty_path))
    output_path = _build_figs6(output_dir, triparty_sheets)
    print(f"Wrote {output_path}")


def _run_all(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sensitivity_path = _sensitivity_payload_path(args.sensitivity_path, output_dir)
    budget_path = (
        Path(args.budget_path)
        if args.budget_path is not None
        else (output_dir / "budget_impact.xlsx")
    )
    copay_path = (
        Path(args.copay_path)
        if args.copay_path is not None
        else (output_dir / "copay_impact.xlsx")
    )
    triparty_path = (
        Path(args.triparty_path)
        if args.triparty_path is not None
        else (output_dir / "triparty_impact.xlsx")
    )

    shared = argparse.Namespace(
        results_glob=args.results_glob,
        results_root=args.results_root,
        output_dir=str(output_dir),
        model_config=args.model_config,
        sensitivity_path=str(sensitivity_path),
        budget_path=str(budget_path),
        copay_path=str(copay_path),
        triparty_path=str(triparty_path),
        price_demand_path=args.price_demand_path,
        price_demand_method=args.price_demand_method,
        price_demand_column=args.price_demand_column,
        price_demand_query=args.price_demand_query,
        price_demand_doses=args.price_demand_doses,
        price_demand_wtp_scale=args.price_demand_wtp_scale,
        copay_person_shares=tuple(args.copay_person_shares),
    )

    _run_tab1(shared)
    _run_tabs1(shared)
    _run_figs1(shared)
    _run_figs2(shared)
    _run_sensitivity(shared)
    _run_tabs3(shared)
    _run_figs3(shared)
    _run_budget(shared)
    _run_figs4(shared)
    _run_copay(shared)
    _run_figs5(shared)
    _run_triparty(shared)
    _run_figs6(shared)
    _run_fig2(shared)
    _run_fig3(shared)


def main() -> None:
    args = _parse_args()
    if args.command == "tab1":
        _run_tab1(args)
        return
    if args.command == "tabs1":
        _run_tabs1(args)
        return
    if args.command == "fig2":
        _run_fig2(args)
        return
    if args.command == "fig3":
        _run_fig3(args)
        return
    if args.command == "price-demand":
        _run_price_demand(args)
        return
    if args.command == "figs1":
        _run_figs1(args)
        return
    if args.command == "figs2":
        _run_figs2(args)
        return
    if args.command == "sensitivity":
        _run_sensitivity(args)
        return
    if args.command == "tabs3":
        _run_tabs3(args)
        return
    if args.command == "figs3":
        _run_figs3(args)
        return
    if args.command == "budget":
        _run_budget(args)
        return
    if args.command == "figs4":
        _run_figs4(args)
        return
    if args.command == "copay":
        _run_copay(args)
        return
    if args.command == "figs5":
        _run_figs5(args)
        return
    if args.command == "triparty":
        _run_triparty(args)
        return
    if args.command == "figs6":
        _run_figs6(args)
        return
    if args.command == "all":
        _run_all(args)
        return
    raise ValueError(f"unsupported command: {args.command}")


if __name__ == "__main__":
    main()
